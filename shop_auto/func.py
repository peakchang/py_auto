import linecache
import random
import threading
import time
import sys
import os
from pathlib import Path
from typing import Optional
import json
import re

import pyautogui as pg
import pyperclip
import pywinauto
import pygetwindow as gw
import clipboard as cb
import openpyxl
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import requests
import winsound as ws
import glob
import random
# import aiohttp
import asyncio



def mainToJisho(driver):
    # 지식쇼핑 안에 들어가서 작업하기!!!!
    mainSearch = searchElement("#MM_SEARCH_FAKE", driver)
    mainSearch[0].click()

    subSearch = searchElement("#query", driver)

    focus_window('NAVER')

    # subSearch[0].send_keys('네이버쇼핑')
    subSearch[0].click()
    keyboard.write(text="네이버쇼핑", delay=0.3)

    searchSubmit = searchElement(".MM_SEARCH_SUBMIT", driver)
    searchSubmit[0].click()

    nShoppingLink = searchElement(".link_name", driver)

    # nShoppingLink[0].click()
    untilEleGone(nShoppingLink[0], ".link_name", driver)




# 함수 시작염

# 상품 들어가서 스크롤 내리고 나오기

def onProductScroll(maxRange, driver):
    
    pg.moveTo(200, 200)
    forCount = 0
    while maxRange > forCount:
        scrollVal = random.randrange(300, 500)
        pg.scroll(-scrollVal)
        wait_float(3, 5)
        forCount += 1
        
    while True:
        wait_float(1.5, 2.3)
        
        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[1])
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            
        try:
            get_shop_list = driver.find_element(by=By.CSS_SELECTOR, value='.mainFilter_option__c4_Lq')
            if get_shop_list is not None:
                return
        except:
            continue

# 지식쇼핑 검색 (2번 해야되니께~~)
def searchJisho(searchKeyword, driver):
    reCount = 0
    while True:
        reCount += 1
        wait_float(0.5,0.9)
        if reCount % 5 == 0:
            driver.refresh()
            pg.press('F5')
        try:
            nShopSearchVar = driver.find_element(by=By.CSS_SELECTOR, value='#sear')
            if(nShopSearchVar):
                break
        except:
            pass
        
        try:
            nShopSearchVar = driver.find_element(by=By.CSS_SELECTOR, value='#input_text')
            if(nShopSearchVar):
                break
        except:
            pass
        try:
            nShopSearchVar = driver.find_element(by=By.CSS_SELECTOR, value='._combineHeader_text_result_8IG-1')
            if(nShopSearchVar):
                break
        except:
            pass
        
    
    driver.execute_script("window.scrollTo(0,0);")
    wait_float(0.5, 1)
    nShopSearchVar.click()

    focus_window("네이버쇼핑")

    wait_float(0.5, 1)
    pg.hotkey('ctrl', 'a')
    wait_float(0.5, 1)
    pg.hotkey('del')
    wait_float(0.5, 1)

    # nShopSearchVar[0].send_keys(searchKeyword)
    # nShopSearchVar.click()
    keyboard.write(text=searchKeyword, delay=0.3)

    wait_float(1.2, 2.5)
    pg.hotkey('enter')
    wait_float(0.5, 1)
    pg.hotkey('enter')

    # 만약 에러나면 검색한 페이지에서 Element가 없어질때까지 loop 추가하자 (네이버 쇼핑 메인은 _lnb_infoscroll-view_1TdpI)

# 엑셀 내 목표클릭과 현재 클릭수 비교해서 배열에 값을 넣음


async def waitPrint(arr, ex, i, chk):
    target_click = int(ex.cell(i, 4).value)
    now_click = ex.cell(i, 5).value

    if now_click is None:
        ex.cell(i, 5).value = 0
        now_click = 0
    now_click = int(now_click)
    if chk == 'Y':
        if now_click < target_click:
            arr.append(i)
    else:
        if now_click >= target_click:
            arr.append(i)
            
async def playAsync_getArr(arr, ex, linkCount, chk):
    try:
        await asyncio.gather(*[waitPrint(arr, ex, i, chk) for i in range(1, linkCount + 1)])
    except:
        pass


# 결과 값(workarr) 을 가지고 해당 인덱스의 엑셀에 1씩 더하기
async def linkExcelPlus(ex, val):
    setVal = ex.cell(val, 5).value
    ex.cell(val, 5).value = setVal + 1


async def playAsync_plusArr(arr, ex):
    try:
        await asyncio.gather(*[linkExcelPlus(ex, val) for val in arr])
    except:
        pass
    

def changeIp():
    try:
        os.system('adb server start')
        client = AdbClient(host="127.0.0.1", port=5037)
        device = client.devices()  # 디바이스 1개
        ondevice = device[0]
        ondevice.shell("input keyevent KEYCODE_POWER")
        ondevice.shell("svc data disable")
        ondevice.shell("settings put global airplane_mode_on 1")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

        ondevice.shell("svc data enable")
        ondevice.shell("settings put global airplane_mode_on 0")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
        time.sleep(3)
        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    except:

        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    return getIp


def changeIpSpeed(driver):
    os.system('adb server start')
    client = AdbClient(host="127.0.0.1", port=5037)
    device = client.devices()  # 디바이스 1개
    ondevice = device[0]
    while True:
        try:
            
            ondevice.shell("input keyevent KEYCODE_POWER")
            ondevice.shell("svc data disable")
            ondevice.shell("settings put global airplane_mode_on 1")
            ondevice.shell(
                "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

            ondevice.shell("svc data enable")
            ondevice.shell("settings put global airplane_mode_on 0")
            ondevice.shell(
                "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
            time.sleep(3)
            while True:
                try:
                    wait_float(0.5, 0.9)
                    getIp = requests.get("http://ip.jsontest.com").json()['ip']
                    if getIp is not None:
                        break
                except:
                    continue
        except:

            while True:
                try:
                    wait_float(0.5, 0.9)
                    getIp = requests.get("http://ip.jsontest.com").json()['ip']
                    if getIp is not None:
                        break
                except:
                    continue
                
        
        driver.get('https://fast.com/ko/')
        searchElement('.speed-results-container', driver)
        time.sleep(3)
        getInternetRapidEle = searchElement('.speed-results-container', driver)
        getInternetRapid = getInternetRapidEle[0].text
        if float(getInternetRapid) < 2.7:
            continue
        else:
            driver.close()
            break
        
    return getIp

def getUaNum():
    with open("./etc/useragent/useragent_all.txt", "r") as f:
        fArr = f.readlines()
        fCount = len(fArr)
        uaSet = random.randrange(0, fCount)
    return uaSet

def searchElement(ele, driver):
    wait_float(0.3, 0.7)
    re_count = 0
    element = ""
    while True:
        re_count += 1
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
            focus_window('chrome')
            pg.press('F5')
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            pass
        

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.3, 0.7)
    return selected_element


def untilEleShow(clickEle, searchEle, driver):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        try:
            btnEle = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is not None:
                return
        except:
            continue


def untilEleGone(clickEle, searchEle, driver):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        
        try:
            btnEle = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is None:
                return
        except:
            return


def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)


def exitApp():
    pg.alert(text='프로그램을 종료합니다.', title='제목입니다.', button='OK')
    sys.exit(0)


def focus_window(winName):
    while True:
        activeName = str(pg.getActiveWindow())
        wait_float(0.3,0.9)
        if winName in activeName:
            return
        winList = pg.getAllWindows()
        # pg.alert(winList)
        wait_float(0.3,0.9)
        for win in winList:
            if winName in win.title:
                pywinauto.application.Application().connect(handle=win._hWnd).top_window().set_focus()
                win.activate()
                wait_float(0.7,1.3)
                break



BASE_DIR = Path(__file__).resolve().parent


def get_secret(
    key: str,
    default_value: Optional[str] = None,
    json_path: str = str(BASE_DIR / "secrets.json"),
):

    with open(json_path) as f:
        secrets = json.loads(f.read())
    try:
        return secrets[key]
    except KeyError:
        if default_value:
            return default_value
        raise EnvironmentError(f"Set the {key} environment variable.")
