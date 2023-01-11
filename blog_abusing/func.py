import random
import threading
import time
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
import sys
import os
from pathlib import Path
from typing import Optional
import requests
import json
import re
import pywinauto
import pyautogui as pg
import pyperclip
import pygetwindow as gw
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.support.select import Select

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import requests
import winsound as ws
import glob
import asyncio
import string


def changeIp():
    os.system('adb server start')
    client = AdbClient(host="127.0.0.1", port=5037)
    device = client.devices()  # 디바이스 1개
    ondevice = device[0]
    ondevice.shell("input keyevent KEYCODE_POWER")
    ondevice.shell("svc data disable")
    ondevice.shell("settings put global airplane_mode_on 1")
    ondevice.shell(
        "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

    ondevice.shell("svc data enable")
    ondevice.shell("settings put global airplane_mode_on 0")
    ondevice.shell(
        "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
    time.sleep(3)
    
    while True:
        try:
            wait_float(0.5, 0.9)
            getIp = requests.get("https://api.ip.pe.kr/json/").json()['ip']
            if getIp is not None:
                break
        except:
            continue
    print(getIp)
    return getIp

def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)
    


def naverLogin_mobile(load_id, load_pass, driver):
    search_bar = searchElement(".sch_ico_aside",driver)
    search_bar[0].click()
    login_btn = searchElement(".ss_profile_wrap",driver)
    login_btn[0].click()

    # 로그인 부분

    focus_window('로그인')
    while True:
        searchElement("#id",driver)

        pyperclip.copy(load_id)
        id_input = driver.find_elements(by=By.CSS_SELECTOR, value="#id")
        id_input[0].click()
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'a')
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'v')
        wait_float(0.4, 0.7)

        pyperclip.copy(load_pass)
        pw_input = driver.find_elements(by=By.CSS_SELECTOR, value="#pw")
        pw_input[0].click()
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'a')
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'v')
        wait_float(0.4, 0.7)
        id_input_value = id_input[0].get_attribute('value')
        if id_input_value:
            pg.hotkey('enter')
            wait_float(0.5, 1.0)
        else:
            continue

        asideChk = 0
        noProblem = ""
        passExit = ""
        while 2 > asideChk:
            try:
                waitAside = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".aside_wrap")))
                if waitAside is not None:
                    noProblem = "on"
                    break
            except:
                asideChk += 1

        if noProblem != "on":
            searchElement("#header",driver)

        try:
            newDevice = driver.find_elements(
                by=By.CSS_SELECTOR, value=".btn_white")
            newDevice[0].click()
        except:
            pass

        try:
            greenBtn = driver.find_elements(
                by=By.CSS_SELECTOR, value=".btn_next")
            greenBtn[0].click()
            passExit = "on"
        except:
            pass

        try:
            protectId = driver.find_elements(
                by=By.CSS_SELECTOR, value=".ico_warning2")
            if protectId:
                return "보호조치"
        except:
            pass

        try:
            sleepId = driver.find_elements(
                by=By.CSS_SELECTOR, value=".warning_v2")
            if sleepId:
                return "휴면아이디"
        except:
            pass

        try:
            unPwd = driver.find_elements(
                by=By.CSS_SELECTOR, value=".error_message")
            if unPwd:
                return "비번틀림"
            # 다시 로그인 어쩌구......
        except:
            pass
        try:
            unPwd = driver.find_elements(
                by=By.CSS_SELECTOR, value=".action_inner")
            if unPwd:
                return "비정상적 활동"
            # 다시 로그인 어쩌구......
        except:
            pass

        if passExit != "on":
            goToMain = searchElement(".ah_close",driver)
            goToMain[0].click()
            wait_float(0.4, 0.7)

        time.sleep(5)
        break
        # 로그인 부분 끝


def searchElement(ele,driver):
    wait_float(0.3, 0.7)
    re_count = 0
    element = ""
    while True:
        re_count += 1
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
            pg.press('F5')
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            pass

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.3, 0.7)
    return selected_element


def focus_window(winName):
    while True:
        activeName = str(pg.getActiveWindow())
        print('에러 11111')
        wait_float(0.3,0.9)
        if winName in activeName:
            return
        
        print('에러 22222')
        winList = pg.getAllWindows()
        # pg.alert(winList)
        wait_float(0.3,0.9)
        print('에러 33333')
        for win in winList:
            print(win.title)
            if winName in win.title:
                pywinauto.application.Application().connect(handle=win._hWnd).top_window().set_focus()
                win.activate()
                wait_float(0.7,1.3)
                break


def mainToPost(driver,blog_list_file,allCount):
    blog_list = blog_list_file.active
    mainSuccess = ''
    while True:
        try:
            searchEle = driver.find_element(by=By.CSS_SELECTOR, value='#MM_SEARCH_FAKE')
            if searchEle:
                break
        except:
            pass
        
        try:
            searchEle = driver.find_element(by=By.CSS_SELECTOR, value='#nx_query')
            if searchEle:
                break
        except:
            pass
        
    focus_window('NAVER')
    searchEle.click()
    wait_float(1.2,2.5)
    
    # 메인 키워드 검색
    mainKeyword = blog_list.cell(allCount, 1).value
    keyboard.write(text=mainKeyword, delay=0.2)
    wait_float(0.5,1.2)
    pg.press('enter')
    
    searchMenu = searchElement('.lst_sch li',driver)
    for menu in searchMenu:
        if menu.text == 'VIEW':
            menu.click()
            break
    
    searchElement('.list_option_filter',driver)
    
    while True:
        wait_float(0.5,1.2)
        postList = driver.find_elements(by=By.CSS_SELECTOR, value='.total_wrap')
        if len(postList) < 90:
            pg.press('end')
        else:
            break
    
    searchTargetLink = blog_list.cell(allCount, 3).value.split('/')[-1]
    for count,post in enumerate(postList):
        getPostHref = post.find_element(by=By.CSS_SELECTOR, value='.api_txt_lines')
        if searchTargetLink in getPostHref.get_attribute('href'):
            driver.execute_script("arguments[0].scrollIntoView();", getPostHref)
            wait_float(0.5,1.2)
            pg.moveTo(300,500)
            pg.scroll(400)
            wait_float(1.2,2.2)
            
            getPostHref.click()
            blog_list.cell(allCount, 4).value = f'메인 - {count}'
            blog_list_file.save('./etc/blog_list.xlsx')
            mainSuccess = 'on'
            break
    
    if mainSuccess == '':
        wait_float(0.5,1.2)
        pg.press('home')
        wait_float(1.2,2.2)
        planService = Select(driver.find_element(by=By.CSS_SELECTOR, value=f'.select_bx'))
        planService.select_by_visible_text('블로그')
        
        searchElement('.select_bx', driver)
        while True:
            wait_float(0.5,1.2)
            postList = driver.find_elements(by=By.CSS_SELECTOR, value='.total_wrap')
            if len(postList) < 300:
                pg.press('end')
            else:
                break
    
        searchTargetLink = blog_list.cell(allCount, 3).value.split('/')[-1]
        for count,post in enumerate(postList):
            getPostHref = post.find_element(by=By.CSS_SELECTOR, value='.api_txt_lines')
            if searchTargetLink in getPostHref.get_attribute('href'):
                driver.execute_script("arguments[0].scrollIntoView();", getPostHref)
                wait_float(0.5,1.2)
                pg.moveTo(300,500)
                pg.scroll(400)
                wait_float(1.2,2.2)
                blog_list.cell(allCount, 4).value = f'제목 - {count}'
                blog_list_file.save('./etc/blog_list.xlsx')
                getPostHref.click()
                mainSuccess = 'on'
                break
        searchElement('.Nservice_item', driver)
        
        