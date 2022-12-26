
import random
import threading
import math
import time
from datetime import date, datetime, timedelta
import sys
import os
from pathlib import Path
from typing import Optional
import pyautogui as pg
import json
import re
from PIL import ImageGrab

import win32com.client
# from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
import pythoncom
import gspread
import openpyxl

from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

"""
주요 변수
nowAction ( write / reply ) / 글쓰는지 댓글인지 체크
nowWriteStatus (optimize / basic ) / 최적화인지 아닌지 체크
allCount : 전체 변수, 글 / 댓글 나누기 위함
writeCount : 글쓰기 변수 nowAction 이 write일때 하나씩 증가, 최적화 아이디인지 판별하기 위함

cafe_id.cell(세로(열), 가로(행)).value
"""

###


def goScript(getDict):
    
    if getDict['getTong'] == 0:
        setTong = 'SK'
    elif getDict['getTong'] == 1:
        setTong = 'KT'
    else:
        setTong = 'LG'
        
    # with open("./result_image/0result.txt", "w") as f:
    #     f.write("start~~~~\n")
    
    
    # 스프레드 시트 열기
    json_file_name = 'ecstatic-magpie-310310-5c58a2ab08ef.json'
    gc = gspread.service_account(filename=json_file_name)
    
    doc = gc.open_by_url('https://docs.google.com/spreadsheets/d/1gWxGWnVPMBN6qDrHglE75Qn5j2Fq9sixEX14mW8qsMo/edit?usp=sharing')
    workSheet = doc.worksheet(setTong)
    
    # 엑셀파일 열기
    excel = win32com.client.Dispatch("Excel.Application", pythoncom.CoInitialize())
    excel.visible = False

    wb = excel.Workbooks.Open(f'{os.getcwd()}/etc/test_ex.xlsx')
    ws = wb.Worksheets['onSheet']
    
    workCreate = openpyxl.load_workbook('./etc/create_link.xlsx')
    sheet = workCreate.get_sheet_by_name('Sheet1')
    
    
    if not getDict['getLine']:
        basicCount = 0
    else:
        basicCount = int(getDict['getLine'])
    
    yogInfoList = workSheet.range(f'B1:H3')
    yogNameList = getArr(yogInfoList, 0)
    yogFeeList = getArr(yogInfoList, 7)
    yogDataList = getArr(yogInfoList, 14)
    
    yogShalList = []
    for i, val in enumerate(yogFeeList):
        sHalVal = math.ceil(val * 0.25) * -1
        yogShalList.append(sHalVal)
        
    setBasicTable(ws, yogNameList, 12, 2)
    setBasicTable(ws, yogFeeList, 13, 3)
    setBasicTable(ws, yogDataList, 16, 6)
    setBasicTable(ws, yogShalList, 14)
    
    if setTong == 'SK':
        createCount = 6
    elif setTong == 'KT':
        createCount = 15
    else:
        createCount = 24
    while True:
        basicCount += 1
        tempVal = workSheet.acell(f'A{basicCount}').value
        if tempVal == 'STOP':
            excel.Quit()
            pg.alert('작업이 완료 되었습니다!')
            break
        if tempVal is not None:
            if '갤럭시' in tempVal or '아이폰' in tempVal:
                deviceName = tempVal
                
                if setTong == 'SK':
                    all_list = workSheet.range(f'B{basicCount}:H{basicCount+9}')
                    capa_list = getArr(all_list, 0, 'ok')
                    fPrice_list = getArr(all_list, 7, 'ok')
                    gongsi_list = getArr(all_list, 14)
                    
                    mnp_ghal_list = getArr(all_list, 42)
                    mnp_shal_list = getArr(all_list, 49)
                    
                    gib_ghal_list = getArr(all_list, 56)
                    gib_shal_list = getArr(all_list, 63)
                    
                else:
                    all_list = workSheet.range(f'B{basicCount}:H{basicCount+8}')
                    capa_list = getArr(all_list, 0, 'ok')
                    fPrice_list = getArr(all_list, 7, 'ok')
                    gongsi_list = getArr(all_list, 14)
                    
                    mnp_ghal_list = getArr(all_list, 35)
                    mnp_shal_list = getArr(all_list, 42)
                    
                    gib_ghal_list = getArr(all_list, 49)
                    gib_shal_list = getArr(all_list, 56)
                
                
                for idx, capa in enumerate(capa_list):
                    
                    ws.cells(5,2).Value = f'{setTong} {deviceName} {capa}'
                    ws.cells(5,12).Value = f'{setTong} {deviceName} {capa}'
                    ws.cells(7,4).Value = fPrice_list[idx]
                    ws.cells(7,15).Value = fPrice_list[idx]
                    setCount = 7
                    for idg, basicFee in enumerate(yogFeeList):
                        ws.cells(5,5).Value = "번호이동 공시지원금 요금제표"
                        ws.cells(setCount+idg,5).Value = gongsi_list[idg]
                        setMnpgHalwon = fPrice_list[idx] - gongsi_list[idg] - mnp_ghal_list[idg]
                        if setMnpgHalwon < 0:
                            setMnpgHalwon = 0
                        setMnpgMonthHal = math.ceil(setMnpgHalwon / 24)
                        ws.cells(setCount+idg,7).Value = setMnpgHalwon
                        ws.cells(setCount+idg,8).Value = setMnpgMonthHal
                        ws.cells(setCount+idg,9).Value = basicFee + setMnpgMonthHal
                        
                        
                        ws.Range(ws.Cells(5,2),ws.Cells(13,9)).Copy()  
                        img = ImageGrab.grabclipboard()
                        imgFile = os.path.join(f'{os.getcwd()}/result_image',f'{setTong}_{pre_val}_{capa}_mnp_gongsi.png')
                        img.save(imgFile)
                        
                        ws.cells(5,15).Value = "번호이동 선택약정 요금제표"
                        setMnpsHalwon = fPrice_list[idx] - mnp_shal_list[idg]
                        if setMnpsHalwon < 0:
                            setMnpsHalwon = 0
                        setMnpsMonthHal = math.ceil(setMnpsHalwon / 24)
                        
                        ws.cells(setCount+idg,17).Value = setMnpsHalwon
                        ws.cells(setCount+idg,18).Value = setMnpsMonthHal
                        ws.cells(setCount+idg,19).Value = basicFee + yogShalList[idg] + setMnpsMonthHal
                        
                        ws.Range(ws.Cells(5,12),ws.Cells(13,19)).Copy()  
                        img = ImageGrab.grabclipboard()
                        imgFile = os.path.join(f'{os.getcwd()}/result_image',f'{setTong}_{pre_val}_{capa}_mnp_sunyak.png')
                        img.save(imgFile)
                        
                    for idg, basicFee in enumerate(yogFeeList):
                        ws.cells(5,5).Value = "기기변경 공시지원금 요금제표"
                        ws.cells(setCount+idg,5).Value = gongsi_list[idg]
                        setgHalwon = fPrice_list[idx] - gongsi_list[idg] - gib_ghal_list[idg]
                        if setgHalwon < 0:
                            setgHalwon = 0
                        setgMonthHal = math.ceil(setgHalwon / 24)
                        ws.cells(setCount+idg,7).Value = setgHalwon
                        ws.cells(setCount+idg,8).Value = setgMonthHal
                        ws.cells(setCount+idg,9).Value = basicFee + setgMonthHal
                        
                        
                        ws.Range(ws.Cells(5,2),ws.Cells(13,9)).Copy()  
                        img = ImageGrab.grabclipboard()
                        imgFile = os.path.join(f'{os.getcwd()}/result_image',f'{setTong}_{pre_val}_{capa}_gib_gongsi.png')
                        img.save(imgFile)

                        
                        ws.cells(5,15).Value = "기기변경 선택약정 요금제표"
                        setsHalwon = fPrice_list[idx] - gib_shal_list[idg]
                        if setsHalwon < 0:
                            setsHalwon = 0
                        setsMonthHal = math.ceil(setsHalwon / 24)
                        
                        ws.cells(setCount+idg,17).Value = setsHalwon
                        ws.cells(setCount+idg,18).Value = setsMonthHal
                        ws.cells(setCount+idg,19).Value = basicFee + yogShalList[idg] + setsMonthHal
                        
                        ws.Range(ws.Cells(5,12),ws.Cells(13,19)).Copy()  
                        img = ImageGrab.grabclipboard()
                        imgFile = os.path.join(f'{os.getcwd()}/result_image',f'{setTong}_{pre_val}_{capa}_gib_sunyak.png')
                        img.save(imgFile)
                    
                    createCount += 1
                    with open("./result_image/0result.txt", "a") as f:
                        f.write(f'{setTong}_{pre_val}_{capa}_gib_gongsi.png,{setTong}_{pre_val}_{capa}_gib_sunyak.png,{setTong}_{pre_val}_{capa}_mnp_gongsi.png,{setTong}_{pre_val}_{capa}_mnp_sunyak.png\n')
                        
                        sheet.cell(2,createCount).value = f'{setTong}_{pre_val}_{capa}_gib_gongsi.png,{setTong}_{pre_val}_{capa}_gib_sunyak.png,{setTong}_{pre_val}_{capa}_mnp_gongsi.png,{setTong}_{pre_val}_{capa}_mnp_sunyak.png'
                        workCreate.save('./etc/create_link.xlsx')
                        
                        
                        f.write(f'{setTong}_{pre_val}_{capa}_gib_gongsi.png\n')
                        f.write(f'{setTong}_{pre_val}_{capa}_gib_sunyak.png\n')
                        f.write(f'{setTong}_{pre_val}_{capa}_mnp_gongsi.png\n')
                        f.write(f'{setTong}_{pre_val}_{capa}_mnp_sunyak.png\n\n')
        pre_val = tempVal

    
    
def make_link():
    # 엑셀파일 열기
    excel = win32com.client.Dispatch("Excel.Application", pythoncom.CoInitialize())
    excel.visible = False
    
    wb = excel.Workbooks.Open(f'{os.getcwd()}/etc/create_link.xlsx')
    ws = wb.Worksheets['Sheet1']
    
    with open("./etc/create_link.txt", "w") as f:
        f.write('생성시작\n\n')
    
    
    plusCount = 0
    while True:
        
        bc = 2+plusCount
        if ws.cells(bc,1).Value is None:
            break
        
        linkText = "http://ts-phone.com/test/update_get.php"
        it_id = f"?it_id={ws.cells(2+plusCount,2).Value}"
        linkText = linkText + it_id
        
        set_tong = f"&it_shop_memo={ws.cells(2+plusCount,3).Value}"
        linkText = linkText + set_tong
        
        set_theme = f"&it_skin={ws.cells(2+plusCount,4).Value}"
        linkText = linkText + set_theme
        # sk_item_list = f"?sk_item_list={ws.cells(2+plusCount,2)}"
        
        sc = 5
        for k in range(3):
            for i in range(9):
                nc = sc + i
                if ws.cells(bc,nc).Value is not None:
                    linkText = linkText + f"&{ws.cells(1,nc).Value}={ws.cells(bc,nc).Value}"
            sc = sc + 9
        pg.alert(linkText)
        
        with open("./etc/create_link.txt", "a") as f:
            f.write(f'{linkText}\n\n')
        plusCount += 1
    
    
            
    pg.alert('종료합니다!!')
    excel.Quit()
    
    
def gogoScript(getDict):
    

    workCreate = openpyxl.load_workbook('./etc/create_link.xlsx')
    sheet = workCreate.get_sheet_by_name('Sheet1')
    if getDict['getTong'] == 0:
        setTong = 'SK'
        exCount = 6
    elif getDict['getTong'] == 1:
        setTong = 'KT'
        exCount = 15
    else:
        setTong = 'LG'
        exCount = 24
        
    # with open("./result_image/0result.txt", "w") as f:
    #     f.write("start~~~~\n")
    
    
    # 스프레드 시트 열기
    json_file_name = 'ecstatic-magpie-310310-5c58a2ab08ef.json'
    gc = gspread.service_account(filename=json_file_name)
    
    doc = gc.open_by_url('https://docs.google.com/spreadsheets/d/1gWxGWnVPMBN6qDrHglE75Qn5j2Fq9sixEX14mW8qsMo/edit?usp=sharing')
    workSheet = doc.worksheet(setTong)
    
    if not getDict['getLine']:
        basicCount = 0
    else:
        basicCount = int(getDict['getLine'])
    
    # yogInfoList = workSheet.range(f'B1:H3')
    # yogNameList = getArr(yogInfoList, 0)
    # yogFeeList = getArr(yogInfoList, 7)
    # yogDataList = getArr(yogInfoList, 14)
    
    # yogShalList = []
    # for i, val in enumerate(yogFeeList):
    #     sHalVal = math.ceil(val * 0.25) * -1
    #     yogShalList.append(sHalVal)

    
    while True:
        basicCount += 1
        tempVal = workSheet.acell(f'A{basicCount}').value
        if tempVal == 'STOP':
            pg.alert('작업이 완료 되었습니다!')
            break
        if tempVal is not None:
            if '갤럭시' in tempVal or '아이폰' in tempVal:
                deviceName = tempVal
                
                if setTong == 'SK':
                    all_list = workSheet.range(f'B{basicCount}:H{basicCount+10}')
                    capa_list = getArr(all_list, 0, 'ok')
                    fPrice_list = getArr(all_list, 7, 'ok')
                    gongsi_list = getArr(all_list, 14)        
                    mnp_ghal_list = getArr(all_list, 42)
                    mnp_shal_list = getArr(all_list, 49)
                    gib_ghal_list = getArr(all_list, 56)
                    gib_shal_list = getArr(all_list, 63)
                else:
                    all_list = workSheet.range(f'B{basicCount}:H{basicCount+9}')
                    capa_list = getArr(all_list, 0, 'ok')
                    fPrice_list = getArr(all_list, 7, 'ok')
                    gongsi_list = getArr(all_list, 14)        
                    mnp_ghal_list = getArr(all_list, 35)
                    mnp_shal_list = getArr(all_list, 42)
                    gib_ghal_list = getArr(all_list, 49)
                    gib_shal_list = getArr(all_list, 56)
                
                for val in fPrice_list:
                    exCount += 1
                    getItemInfo = f"{val}|{','.join(str(_) for _ in gongsi_list)}|{','.join(str(_) for _ in gib_ghal_list)}|{','.join(str(_) for _ in gib_shal_list)}|{','.join(str(_) for _ in mnp_ghal_list)}|{','.join(str(_) for _ in mnp_shal_list)}"
                    
                    sheet.cell(2,exCount).value = getItemInfo
                    
                    workCreate.save('./etc/create_link.xlsx')



# ******************************************************** 계산기 기준!!!!!


def calculScript(getDict):
    
    goTongArr = getDict['goTong'].split(',')
    getLineArr = getDict['getLine'].split(',')
    
    with open("./etc/min_price.txt", "w") as f:
        f.write("start~~~~\n")
    
    try:
        int(getLineArr[0])
    except:
        pg.alert('라인을 입력해주세요! 종료합니다!')
        return
    
    # 추후 중저가폰 할때 정하기
    highendDevice = ",0,4,5,5"
    # LogDevice = 
    # kidsDevice = 
    
    # 엑셀 열기
    workCreate = openpyxl.load_workbook('./etc/create_calcul_link.xlsx')
    sheet = workCreate.get_sheet_by_name('Sheet1')
    
    
    for ii in range(30):
        if ii == 0:
            continue
        sheet.cell(2, ii).value = ""
        
    
    
    # 스프레드 시트 열기
    json_file_name = 'ecstatic-magpie-310310-5c58a2ab08ef.json'
    gc = gspread.service_account(filename=json_file_name)
    
    doc = gc.open_by_url('https://docs.google.com/spreadsheets/d/1gWxGWnVPMBN6qDrHglE75Qn5j2Fq9sixEX14mW8qsMo/edit?usp=sharing')
    
    for idx, nowTong in enumerate(goTongArr):
        workSheet = doc.worksheet(nowTong)
        basicCount = int(getLineArr[idx])
        
        
        while True:
            basicCount += 1
            tempVal = workSheet.acell(f'A{basicCount}').value
            if tempVal is not None:
                if '갤럭시' in tempVal or '아이폰' in tempVal:
                    deviceName = tempVal
                    
                    if nowTong == 'SK':
                        all_list = workSheet.range(f'B{basicCount}:H{basicCount+13}')
                        capa_list = getArrToStr(all_list, 0, 'ok')
                        fPrice_list = getArrToStr(all_list, 7, 'ok')
                        gongsi_list = getArrToStr(all_list, 14)
                        mnp_ghal_list = getArrToStr(all_list, 42)
                        mnp_shal_list = getArrToStr(all_list, 49)        
                        gib_ghal_list = getArrToStr(all_list, 56)
                        gib_shal_list = getArrToStr(all_list, 63)
                        
                        sheet.cell(2,5).value = fPrice_list
                        sheet.cell(2,6).value = capa_list
                        sheet.cell(2,7).value = gongsi_list
                        sheet.cell(2,8).value = mnp_ghal_list
                        sheet.cell(2,9).value = mnp_shal_list
                        sheet.cell(2,10).value = gib_ghal_list
                        sheet.cell(2,11).value = gib_shal_list
                        workCreate.save('./etc/create_calcul_link.xlsx')
                        priceList1 = getArr(all_list, 70)
                        priceList2 = getArr(all_list, 77)
                        priceList3 = getArr(all_list, 84)
                        priceList4 = getArr(all_list, 91)
                        
                        price_list = priceList1 + priceList2 + priceList3 + priceList4
                        minPrice = min(price_list)
                        with open("./etc/min_price.txt", "a") as f:
                            f.write(f"SK 최저가 : \n{minPrice}\n")
                    elif nowTong == 'KT':
                        all_list = workSheet.range(f'B{basicCount}:H{basicCount+12}')
                        capa_list = getArrToStr(all_list, 0, 'ok')
                        fPrice_list = getArrToStr(all_list, 7, 'ok')
                        gongsi_list = getArrToStr(all_list, 14)
                        mnp_ghal_list = getArrToStr(all_list, 35)
                        mnp_shal_list = getArrToStr(all_list, 42)
                        gib_ghal_list = getArrToStr(all_list, 49)
                        gib_shal_list = getArrToStr(all_list, 56)
                        
                        sheet.cell(2,12).value = fPrice_list
                        sheet.cell(2,13).value = capa_list
                        sheet.cell(2,14).value = gongsi_list
                        sheet.cell(2,15).value = mnp_ghal_list
                        sheet.cell(2,16).value = mnp_shal_list
                        sheet.cell(2,17).value = gib_ghal_list
                        sheet.cell(2,18).value = gib_shal_list
                        workCreate.save('./etc/create_calcul_link.xlsx')
                        priceList1 = getArr(all_list, 63)
                        priceList2 = getArr(all_list, 70)
                        priceList3 = getArr(all_list, 77)
                        priceList4 = getArr(all_list, 84)
                        
                        price_list = priceList1 + priceList2 + priceList3 + priceList4
                        
                        minPrice = min(price_list)
                        with open("./etc/min_price.txt", "a") as f:
                            f.write(f"KT 최저가 : \n{minPrice}\n")
                                
                    else:
                        all_list = workSheet.range(f'B{basicCount}:H{basicCount+12}')
                        capa_list = getArrToStr(all_list, 0, 'ok')
                        fPrice_list = getArrToStr(all_list, 7, 'ok')
                        gongsi_list = getArrToStr(all_list, 14)
                        
                        mnp_ghal_list = getArrToStr(all_list, 35)
                        mnp_shal_list = getArrToStr(all_list, 42)
                        
                        gib_ghal_list = getArrToStr(all_list, 49)
                        gib_shal_list = getArrToStr(all_list, 56)
                        
                        priceList1 = getArr(all_list, 63)
                        priceList2 = getArr(all_list, 70)
                        priceList3 = getArr(all_list, 77)
                        priceList4 = getArr(all_list, 84)
                        
                        sheet.cell(2,19).value = fPrice_list
                        sheet.cell(2,20).value = capa_list
                        sheet.cell(2,21).value = gongsi_list
                        sheet.cell(2,22).value = mnp_ghal_list
                        sheet.cell(2,23).value = mnp_shal_list
                        sheet.cell(2,24).value = gib_ghal_list
                        sheet.cell(2,25).value = gib_shal_list
                        workCreate.save('./etc/create_calcul_link.xlsx')
                        price_list = priceList1 + priceList2 + priceList3 + priceList4
                        
                        minPrice = min(price_list)
                        with open("./etc/min_price.txt", "a") as f:
                            f.write(f"LG 최저가 : \n{minPrice}\n")
                        
                                
                    
                    break
    
    pg.alert('완료 되었습니다!')
    
    
def make_link_calcul():
    # 엑셀파일 열기

    workCreate = openpyxl.load_workbook('./etc/create_calcul_link.xlsx')
    sheet = workCreate.get_sheet_by_name('Sheet1')
    with open("./etc/create_calcul_link.txt", "w") as f:
        f.write('생성시작\n\n')
    
    linkText = "http://ts-phone.com/test/update_calcul.php"
    onCount = 0
    while True:
        onCount += 1
        nameValue = sheet.cell(1,onCount).value
        valValue = sheet.cell(2,onCount).value
        if nameValue is None:
            break
        
        if onCount == 1:
            addLink = f"?{nameValue}={valValue}"
        else:
            addLink = f"&{nameValue}={valValue}"
        linkText = linkText + addLink
        
    with open("./etc/create_calcul_link.txt", "w") as f:
        f.write(f'{linkText}')
    
    pg.alert('종료합니다!!')



def getGongsi(getDict):
    
    with open(f'./etc/useragent_all.txt', 'r') as f:
        userAgentList = f.readlines()
    ranVal = random.randrange(0,len(userAgentList))
    ua_data = userAgentList[ranVal].replace('\n','')
    pg.alert(ua_data)

    options = Options()
    user_agent = ua_data
    options.add_argument('user-agent=' + user_agent)
    global driver
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(chrome_options=options, service=service)
    
    driver.get('http://www.smartchoice.or.kr/smc/mobile/dantongTelList.do')
    
    nowTong = getDict['getTong']
    
    # 엑셀 열기
    workCreate = openpyxl.load_workbook('./etc/get_gongsi.xlsx')
    sheet = workCreate.get_sheet_by_name(nowTong)
    
    
    
    sheet.cell(2,11).value
    
    getCount = 1
    while True:
        preResult = ''
        getCount += 1
        pServiceName = sheet.cell(getCount,1).value
        if pServiceName is None:
            pg.alert('완료 되었습니다!')
            driver.quit()
            break
        dMauName = sheet.cell(getCount,2).value
        deviceName = sheet.cell(getCount,3).value

        danCompany = Select(driver.find_element(by=By.CSS_SELECTOR, value='#dan_Company'))
        danCompany.select_by_visible_text(nowTong)
        time.sleep(2)

        planService = Select(driver.find_element(by=By.CSS_SELECTOR, value=f'#plan{nowTong}Service'))
        planService.select_by_visible_text(pServiceName)
        time.sleep(2)
        
        danMau = Select(driver.find_element(by=By.CSS_SELECTOR, value=f'#dan_Mau'))
        danMau.select_by_visible_text(dMauName)
        time.sleep(2)
        
        deviceModalBtn = driver.find_element(by=By.CSS_SELECTOR, value=f'#product_btn')
        deviceModalBtn.click()
        time.sleep(3)
        
        
        popupList = driver.find_elements(by=By.CSS_SELECTOR, value=f'.popcontents')
        for searchPop in popupList:
            try:
                if '휴대폰' in searchPop.text:
                    setPop = searchPop
                    time.sleep(2)
                    break
            except:
                pass
        
        deviceList = setPop.find_elements(by=By.CSS_SELECTOR, value=f'.monthlyValue')

        for val in deviceList:
            if val.text == deviceName:
                val.click()
                break

        time.sleep(2)
        
        mothlybtn = driver.find_elements(by=By.CSS_SELECTOR, value='.mothlybtn')
        mothlybtn[1].click()
        time.sleep(2)
        
        yogCount = 3
        while True:
            print('에러예상 0000')
            yogCount += 1
            getYogName = sheet.cell(1,yogCount).value
            if getYogName is None:
                break
            
            print('에러예상 1111')
            planBtn = driver.find_element(by=By.CSS_SELECTOR, value='#plan_btn')
            planBtn.click()
            time.sleep(1)
            
            print('에러예상 2222')
            
            
            yogListPopUp = driver.find_element(by=By.CSS_SELECTOR, value=f'.selectPopup')
            time.sleep(2)
            yogList = yogListPopUp.find_elements(by=By.CSS_SELECTOR, value=f'.monthlyValue')
            
            for val in yogList:
                if val.text == getYogName:
                    val.click()
                    break
            time.sleep(1)
            print('에러예상 3333')
            mothlybtn = driver.find_elements(by=By.CSS_SELECTOR, value='.mothlybtn')
            mothlybtn[1].click()
            time.sleep(1)
            print('에러예상 4444')
            searchBtn = driver.find_elements(by=By.CSS_SELECTOR, value='.btn_wrap.item2.mt10 a')
            searchBtn[1].click()
            time.sleep(2)
            
            sameCount = 0
            while True:
                if sameCount > 15:
                    break
                findResult = driver.find_elements(by=By.CSS_SELECTOR, value='.findResult td')
                try:
                    getResult = findResult[2].text
                except:
                    continue
                if getResult == preResult:
                    sameCount += 1
                    continue
                else:
                    break
                
            sheet.cell(getCount,yogCount).value = findResult[2].text
            preResult = findResult[2].text
            workCreate.save('./etc/get_gongsi.xlsx')









    
def getArr(setList, setNum, ok=''):
    temp_list = setList[setNum:setNum+7]
    temp_arr = []
    for val in temp_list:
        if not val.value:
            if ok:
                continue
            else:
                temp_arr.append(0)
        else:
            try:
                setVal = int(val.value)
            except:
                setVal = val.value
            temp_arr.append(setVal)
    return temp_arr

def getArrToStr(setList, setNum, ok=''):
    temp_list = setList[setNum:setNum+7]
    temp_arr = []
    for val in temp_list:
        if not val.value:
            if ok:
                continue
            else:
                temp_arr.append(0)
        else:
            try:
                setVal = int(val.value)
            except:
                setVal = val.value
            temp_arr.append(setVal)
    
    temp_str = ','.join(str(_) for _ in temp_arr)
    return temp_str
        
def setBasicTable(ex, list, scount, gcount = ''):
    startCount = 7
    for i, val in enumerate(list):
        ex.cells(startCount+i,scount).Value = val
        if gcount:
            ex.cells(startCount+i,gcount).Value = val
        