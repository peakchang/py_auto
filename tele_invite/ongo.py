import random
import threading
import time
from datetime import datetime, timedelta, date
import sys
import os
from pathlib import Path
import json
import re
import pyautogui as pg
import pyperclip
import pygetwindow as gw
import clipboard as cb
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import WebDriverException

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
from tkinter import *
import tkinter
from tkinter import ttk
import requests
import winsound as ws
import glob
import asyncio

import getmac
import getpass

import shutil
import winsound as sd

import httpimport



def goScript(getDict):
    
    showLog = 1
    showAlert = 0
    
    if showAlert:
        pg.alert('보자보자')

    
    
    
    try:
        with open(f'./auth.txt', 'r', encoding='UTF8') as r:
            get_auth = re.sub(r'[/s]', '', r.read())
    except:
        with open(f'./auth.txt', 'r') as r:
            get_auth = re.sub(r'[/s]', '', r.read())
    
    get_mac = getmac.get_mac_address()
    
    webhook_url = "https://adpeak.kr/telework/gethook"
    data = {'get_auth' : get_auth, 'get_mac' : get_mac}
    requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)
    r = requests.post(webhook_url, data=json.dumps(data), headers={'Content-Type' : 'application/json'}, verify=False)
    wh_result = r.json()
    
    if wh_result['get_status'] == 'no':
        pg.alert('인증에 실패하였습니다. 관리자에게 문의해주세요')
        sys.exit(0)
    elif wh_result['get_status'] == 'retry':
        pg.alert('등록 되었습니다. 다시 시도 해주세요')
        sys.exit(0)
    elif wh_result['get_status'] == 'ok':
        func_url = wh_result['hidden_link']

    with httpimport.remote_repo(func_url):
        import chk_tele
    
    
    pg.FAILSAFE = False
    pg.alert('시작합니당')
    

    pcUser = getpass.getuser()
    authList = load_workbook('./auth_list.xlsx')
    authSheet = authList.active
    
    dbList = load_workbook('./db_list.xlsx')
    dbSheet = dbList.active
    chkInnerUserText = ['전까지','어제','일주일','최근','오늘','온라인', '방금']
        
    noMoreDb = ''
    authCount = 0
    allCount = 0
    while True:
        allCount += 1
        allCountChk = authSheet.cell(allCount, 1).value
        if allCountChk is None:
            break
    allCount = allCount - 2
    while True:
        try:
            
            while True:
                authCount += 1
                authChk = authSheet.cell(authCount, 6).value
                if authChk is None:
                    break
            
            if showLog:
                print(f'{authCount}번째 아이디 작업 시작')
                
            endChk = authSheet.cell(authCount, 1).value
            if endChk is None:
                pg.alert('작업이 완료 되었습니다.')
                break
            elif noMoreDb == 'on':
                pg.alert('DB가 소진 되었습니다. 작업을 종료합니다')
                break
            
            today = datetime.today()
            todayStr = today.strftime("%y/%m/%d")
            
            profileNum = authSheet.cell(authCount,1).value
            workType = authSheet.cell(authCount,5).value
            if profileNum is None:
                break
            
            profileStatus = authSheet.cell(authCount,3).value
            if profileStatus is None or "X" in profileStatus or 'x' in profileStatus:
                authSheet.cell(authCount, 6).value = '미인증 패스'
                authList.save('./auth_list.xlsx')
                continue
            if "X" not in profileStatus and "x" not in profileStatus:
                
                print(f"총 {allCount}개 중 {authCount - 1}번째 {profileNum}번 아이디 진행중입니다...")
                
                
                options = Options()
                user_data = f'C:\\Users\\{pcUser}\\AppData\\Local\\Google\\Chrome\\User Data\\default'
                service = Service(ChromeDriverManager().install())
                options.add_argument(f"user-data-dir={user_data}")
                options.add_argument(f'--profile-directory=Profile {profileNum}')
                driver = webdriver.Chrome(service=service, chrome_options=options)
                driver.get('https://web.telegram.org/z/')
                driver.set_window_size(1600, 800)
                driver.set_window_position(0,0)
                
                if showLog:
                    print('텔레그램 접속')
                
                
                
                fore = pg.getActiveWindow()
                # print(fore.title)   # 활성화된 창의 제목 정보
                # print(fore.size)    # 활성화된 창의 사이즈
                # print(fore.left, fore.top, fore.right, fore.bottom) # 좌표정보
                
                
                chk_tele.wait_float(2.5,3.2)
                
                
                    
                notAuth = ''
                okAuth = ''
                
                errCount = 0
                while True:
                    errCount += 1
                    if errCount > 5:
                        errCount = 0
                        pg.press('F5')
                    if showLog:
                        print('아이디가 짤렸는지 안짤렸는디 최초 검증!!')
                    try:
                        okAuth = driver.find_element(by=By.CSS_SELECTOR, value='#MiddleColumn')
                        if okAuth:
                            break
                    except:
                        pass
                    
                    try:
                        notAuth = driver.find_element(by=By.CSS_SELECTOR, value='#auth-qr-form')
                        if notAuth:
                            break
                    except:
                        pass
                
                if showLog:
                    print(f'okAuth : {okAuth} 인증 정상!! // notAuth : {notAuth} 인증 실패!!')
                    
                if notAuth != '':
                    pg.alert('인증이 안되어있옹')
                    authSheet.cell(authCount,3).value = '인증XX'
                    authList.save('./auth_list.xlsx')
                    driver.quit()
                    continue
                
                
                # 만약 현재 영어 버전일경우 한글 버전으로 변경!!
                chkError = chk_tele.changeToKorean(driver, fore)
                if chkError == 'onerror':
                    authCount = authCount - 1
                    continue
                
                if showLog:
                    print(f'한글 버전 체크 완료!! 준비 끝!!')
                
                if getDict['add_addr_val'] or getDict['add_addr_only']:
                    
                    if getDict['add_addr_only']:
                        chk_tele.goToMain(driver, fore)
                        if showLog:
                            print('메인 가기 완료')
                        tempPhNum = '010-2190-2197'
                        notMb = ''
                        finChk = ''
                        
                        if showAlert:
                            pg.alert('연락처 추가 대기!!!!')
                        chk_tele.addAddr(driver,fore,tempPhNum,getDict['max_addr_count'])
                        if showAlert:
                            pg.alert('연락처 추가 완료!!!!')
                        
                        while True:
                            chk_tele.wait_float(2.7,3.5)
                            # print('친추 완료! 모달창 떠있으면 가입한 회원 아님 / 안떠있으면 체크!')
                            try:
                                notMb = driver.find_element(by=By.CSS_SELECTOR, value='.NewContactModal__new-contact')
                                if notMb:
                                    break
                            except:
                                pass
                            
                            try:
                                findMb = driver.find_element(by=By.CSS_SELECTOR, value='.MiddleHeader')
                                if findMb:
                                    break
                            except:
                                pass
                        if showAlert:
                            pg.alert('연락처 추가 확인')
                            
                        if notMb:
                            chk_tele.wait_float(1.2,1.9)
                            authSheet.cell(authCount, 6).value = f"연락처 추가 기능 짤림"
                            authList.save('./auth_list.xlsx')
                            driver.quit()
                            pg.click(fore.left+500,fore.top+300)
                            continue
                        
                        else:
                            refreshCount = 0
                            setPhNum = re.sub(r'[^0-9]', '', str(tempPhNum))
                            while True:
                                refreshCount += 1
                                chk_tele.wait_float(1.2,1.9)
                                if showLog:
                                    print(refreshCount)
                                if refreshCount % 3 == 0:
                                    if refreshCount % 6 == 0:
                                        if showAlert:
                                            pg.alert('user-status가 안뜰경우 회원 클릭')
                                        chk_tele.wrongUserWork(driver,fore,setPhNum,1)
                                    else:
                                        if showAlert:
                                            pg.alert('user-status가 안떠서 새로고침')
                                        pg.press('F5')
                                        
                                userStatus = chk_tele.searchWaitElement('.MiddleHeader .user-status', driver)
                                userStatusText = re.sub(r'[\s]', '', userStatus[0].text)
                                if userStatusText:
                                    break
                                
                            # 연락처 삭제 준비, 삭제 아이콘 나오게
                            chk_tele.searchAndClick('.icon-delete', '.tools button', driver, 1, '.MiddleHeader .fullName')
                            
                            chk_tele.openModalAndDelete('회원 정보',driver)
                            # # 연락처 삭제 모달창 띄우기
                            # chk_tele.searchAndClick('.Modal', '.destructive', driver)
                                
                            # # 연락처 삭제 완료
                            # chk_tele.searchTextAndClick('회원 정보', '.Modal .confirm-dialog-button.default.danger.text', driver)
                        
                    # DB 카운트 ID값 미 기재된 라인 count 찾기!!
                    dbCount = 0
                    while True:
                        dbCount += 1
                        dbId = dbSheet.cell(dbCount,1).value
                        if dbId is None:
                            break
                    # 준비 완료!! 사람 추가 반복하자!!
                    for i in range(int(getDict['add_addr_count'])):
                        chk_tele.goToMain(driver, fore)
                        if showLog:
                            print('메인 가기 완료')
                        notMb = ''
                        finChk = ''
                        dbLine = dbCount + i
                        getPhNum = dbSheet.cell(dbLine,4).value
                        if getPhNum is None:
                            noMoreDb = 'on'
                            break
                        
                        getPhNum = chk_tele.changePhNum(getPhNum)

                        dbSheet.cell(dbLine,1).value = profileNum
                        dbList.save('./db_list.xlsx')
                        if showLog:
                            print('연락처 추가하기! 모달창 키고 번호 입력!')
                            
                        maxAddrFull = ''
                        maxAddrFull = chk_tele.addAddr(driver,fore,getPhNum,getDict['max_addr_count'])
                        
                        if maxAddrFull == 'on':
                            if showAlert:
                                pg.alert('연락처 꽉참 다음으로')
                            break
                            
                        #친추 완료! 모달창 떠있으면 가입한 회원 아님 / 안떠있으면 체크!
                        
                        
                        while True:
                            chk_tele.wait_float(2.7,3.5)
                            if showLog:
                                print('친추 완료! 모달창 떠있으면 가입한 회원 아님 / 안떠있으면 체크!')
                            try:
                                notMb = driver.find_element(by=By.CSS_SELECTOR, value='.NewContactModal__new-contact')
                                if notMb:
                                    break
                            except:
                                pass
                            
                            try:
                                findMb = driver.find_element(by=By.CSS_SELECTOR, value='.MiddleHeader')
                                if findMb:
                                    break
                            except:
                                pass
                            
                        if notMb:
                            chk_tele.wait_float(1.2,1.9)
                            dbSheet.cell(dbLine,5).value = 'V'
                            dbList.save('./db_list.xlsx')
                            pg.click(fore.left+500,fore.top+300)
                            continue
                        
                        if showAlert:
                            pg.alert('일단 가입된 아이디임')
                        
                        if showLog:
                            print('일단 가입된 아이디임')
                        
                        refreshCount = 0
                        setPhNum = re.sub(r'[^0-9]', '', str(getPhNum))
                        while True:
                            refreshCount += 1
                            chk_tele.wait_float(1.2,1.9)
                            if showLog:
                                print(refreshCount)
                            if refreshCount % 3 == 0:
                                if refreshCount % 6 == 0:
                                    refreshCount = 0
                                    if showAlert:
                                        pg.alert('user-status가 안뜰경우 회원 클릭')
                                    if showLog:
                                        print('user-status가 안뜰경우 회원 클릭')
                                    chk_tele.wrongUserWork(driver,fore,setPhNum,1)
                                else:
                                    if showAlert:
                                        pg.alert('새로고침 하기~')
                                    if showLog:
                                        print('새로고침 하기~')
                                    pg.press('F5')
                                    
                            userStatus = chk_tele.searchWaitElement('.MiddleHeader .user-status', driver)
                            userStatusText = re.sub(r'[\s]', '', userStatus[0].text)
                            if userStatusText:
                                break
                        
                        if showAlert:
                            pg.alert(f'상단 텍스트 {userStatus[0].text} 확인하기')
                        
                        
                        for chkText in chkInnerUserText:
                            if chkText in userStatusText:
                                chk_tele.wait_float(1.2,1.9)
                                dbSheet.cell(dbLine,7).value = 'V'
                                dbList.save('./db_list.xlsx')
                                finChk = 'ok'
                                if getDict['add_addr_only']:
                                    
                                    if showAlert:
                                        pg.alert(f'전부 정상인거 확인! 삭제 시작')
                                        
                                    # 연락처 삭제 준비, 삭제 아이콘 나오게
                                    chk_tele.searchAndClick('.icon-delete', '.tools button', driver, 1, '.MiddleHeader .fullName')
                                    
                                    chk_tele.openModalAndDelete('회원 정보',driver)
                                
                                    # # 연락처 삭제 모달창 띄우기
                                    # chk_tele.searchAndClick('.Modal', '.destructive', driver)
                                    
                                    # # 연락처 삭제 완료
                                    # chk_tele.searchTextAndClick('회원 정보', '.Modal .confirm-dialog-button.default.danger.text', driver)
                                    continue
                        
                        oldUser = ''
                        if finChk == '':
                            chk_tele.wait_float(1.2,1.9)
                            if "오래됨" in userStatusText or "마지막" in userStatusText:
                                oldUser = 'on'
                            else:
                                try:
                                    minus_date = int(getDict['search_day'])
                                    chkCompare = chk_tele.compareDate(userStatusText,minus_date)
                                except:
                                    for i in range(3):
                                        fr = 1600    # range : 37 ~ 32767
                                        du = 500     # 1000 ms ==1second
                                        sd.Beep(fr, du)
                                    pg.alert('초대 변수 예외 에러 발생! 현재 화면 캡쳐해서, 관리자에게 문의 주세요!')
                                
                            if oldUser == '' and chkCompare:
                                dbSheet.cell(dbLine,7).value = 'V'
                                dbList.save('./db_list.xlsx')
                                if getDict['add_addr_only']:
                                    # 연락처 삭제 준비, 삭제 아이콘 나오게
                                    chk_tele.searchAndClick('.icon-delete', '.tools button', driver, 1, '.MiddleHeader .fullName')

                                    chk_tele.openModalAndDelete('회원 정보',driver)
                                    # # 연락처 삭제 모달창 띄우기
                                    # chk_tele.searchAndClick('.Modal', '.destructive', driver)
                                    
                                    # # 연락처 삭제 완료
                                    # chk_tele.searchTextAndClick('회원 정보', '.Modal .confirm-dialog-button.default.danger.text', driver)
                            else:
                                dbSheet.cell(dbLine,6).value = 'V'
                                dbList.save('./db_list.xlsx')
                                
                                # 연락처 삭제 준비, 삭제 아이콘 나오게
                                chk_tele.searchAndClick('.icon-delete', '.tools button', driver, 1, '.MiddleHeader .fullName')
                                
                                chk_tele.openModalAndDelete('회원 정보',driver)
                                
                                # # 연락처 삭제 모달창 띄우기
                                # chk_tele.searchAndClick('.Modal', '.destructive', driver)
                                    
                                # # 연락처 삭제 완료
                                # chk_tele.searchTextAndClick('회원 정보', '.Modal .confirm-dialog-button.default.danger.text', driver)
                                    
                if not getDict['join_group_val'] or getDict['add_addr_only']:
                    authSheet.cell(authCount, 6).value = f"{todayStr} 작업 완료"
                    authList.save('./auth_list.xlsx')
                    driver.quit()
                    continue
                
                   
                ################## 아이디 추가 작업 끝 그룹에 추가 시작!!
                
                if getDict['join_group_val'] and not getDict['add_addr_only']:
                    
                    getChatRoomName = authSheet.cell(authCount,4).value.strip()
                    saveGroupType = ""
                    exceptUserList = []
                    
                    while True:
                        
                         
                        
                        # 그룹 클릭 (그룹명 찾아서 클릭 / 채팅방 클릭)
                        noMatchCount = 0
                        while True:
                            noMatchCount += 1
                            chk_tele.goToMain(driver, fore)
                            if showLog:
                                print("그룹 클릭 (그룹명 찾아서 클릭 / 채팅방 클릭)")
                            chk_tele.wait_float(0.9,1.5)
                            try:
                                nowChatRoom = driver.find_element(by=By.CSS_SELECTOR, value='.MiddleHeader .ChatInfo .fullName')
                                if getChatRoomName in nowChatRoom.text:
                                    break
                            except:
                                if noMatchCount > 5:
                                    pg.alert('그룹명을 찾을수 없습니다. 확인 후 다시 실행 해주세요')
                                    sys.exit(0)
                                else:
                                    pass
                            
                            try:
                                chk_tele.wait_float(0.5,0.9)
                                chatList = chk_tele.searchWaitElement('.chat-list .ListItem', driver)
                                for chatRoom in chatList:
                                    chk_tele.wait_float(0.2,0.5)
                                    if getChatRoomName in chatRoom.text:
                                        chk_tele.wait_float(0.5,0.9)
                                        chatRoom.click()
                                        break
                            except:
                                pass
                        if showAlert:
                            pg.alert('그룹 클릭 완료~')
                            
                        chk_tele.wait_float(0.5,0.9)
                        
                        # 상단 그룹이름 클릭(우측 그룹 정보 나올때까지)
                        while True:
                            if showLog:
                                print('그룹 관리 열기')
                            try:
                                chk_tele.wait_float(0.9,1.2)
                                groupInfo = driver.find_element(by=By.CSS_SELECTOR, value='.chat-info-wrapper .ChatInfo')
                                groupStatus = groupInfo.find_element(by=By.CSS_SELECTOR, value='.group-status')
                                preGroupCount = re.sub(r'[^0-9]', '', groupStatus.text)
                                groupInfo.click()
                            except:
                                pass
                            
                            try:
                                chk_tele.wait_float(0.9,1.2)
                                ProfilePhoto = driver.find_element(by=By.CSS_SELECTOR, value='.ProfilePhoto')
                                if ProfilePhoto:
                                    break
                            except:
                                pg.press('F5')
                                pass
                        if showAlert:
                            pg.alert('그룹 상단 클릭 완료~ 우측에 프로필 나오면 성공')
                            
                        if workType == '관리자추가':
                            
                            # 그룹 정보 우상단 연필 클릭
                            # print('그룹 툴 열기')
                            chk_tele.searchAndClick('.AvatarEditable', '.tools button', driver)

                            chk_tele.wait_float(0.5,0.9)
                            
                            groupMenu = chk_tele.searchWaitElement('.Management .ListItem', driver)
                            # 그룹 > 수정 > 관리자 클릭
                            while True:
                                # print("그룹 > 수정 > 관리자 클릭")
                                try:
                                    chk_tele.wait_float(0.9,1.2)
                                    tools = driver.find_elements(by=By.CSS_SELECTOR, value='.tools button')
                                    tools[0].click()
                                except:
                                    pass
                                
                                try:
                                    chk_tele.wait_float(0.9,1.2)
                                    groupMenu = chk_tele.searchWaitElement('.Management .ListItem', driver)
                                    for menu in groupMenu:
                                        if "관리자" in menu.text:
                                            menu.click()
                                            
                                    chk_tele.wait_float(0.9,1.2)
                                    managerAddBtn = driver.find_element(by=By.CSS_SELECTOR, value='.FloatingActionButton.revealed')
                                    managerAddBtnText = managerAddBtn.get_attribute('title')
                                    if "추가" in managerAddBtnText:
                                        managerAddBtn.click()
                                        break
                                except:
                                    pass
                            
                            
                            # 관리자 추가 > 010 검색 > 010 번호 가진사람 클릭
                            while True:
                                chk_tele.wait_float(0.5,0.9)
                                # print("관리자 추가 > 010 검색 > 010 번호 가진사람 클릭")
                                findUser = ''
                                try:
                                    chk_tele.wait_float(0.5,0.9)
                                    searchAddMgInput = chk_tele.searchWaitElement('.Management__filter .form-control', driver)
                                    getVal = searchAddMgInput[0].get_attribute('value')
                                    if getVal:
                                        pass
                                    else:
                                        searchAddMgInput[0].send_keys("010")
                                    
                                    chk_tele.wait_float(2.1,2.9)
                                    searchUserNameList = driver.find_elements(by=By.CSS_SELECTOR, value='.Management .picker-list .ListItem .ChatInfo .fullName')
                                    for userName in searchUserNameList:
                                        if userName.text[0:3] == '010' or userName.text[0:2] == '10':
                                            userName.click()
                                            setUserName = re.sub(r'[^0-9]', '', userName.text)
                                            findUser = 'on'
                                            break
                                        
                                    chk_tele.wait_float(0.5,0.9)
                                    
                                    if findUser == 'on' and setUserName:
                                        menegerOkBtn = chk_tele.searchWaitElement('.Management .FloatingActionButton', driver)
                                        if menegerOkBtn[0]:
                                            break
                                    elif findUser == '':
                                        break
                                except:
                                    pass
                                
                                try:
                                    nothingFound = driver.find_element(by=By.CSS_SELECTOR, value='.Management .NothingFound')
                                    if nothingFound:
                                        findUser = ''
                                        break
                                except:
                                    pass
                                
                            if findUser == '':
                                # print('더이상 찾을 회원이 없음')
                                authSheet.cell(authCount, 6).value = f"{todayStr} 작업 완료"
                                authList.save('./auth_list.xlsx')
                                break
                            
                            
                            wrongUser = ''
                            while True:
                                # print("추가된 사람 관리자 승격")
                                chk_tele.wait_float(1.2,1.9)
                                
                                # 초대 거절 설정해놓은 회원 거르기
                                try:
                                    # print('초대 거절 확인')
                                    refuseUserChkModal = driver.find_element(by=By.CSS_SELECTOR, value='.modal-title')
                                    if refuseUserChkModal or 'wrong' in refuseUserChkModal.text:
                                        # print('초대 거절 하신 회원이군요~~~')
                                        chk_tele.wait_float(0.3,0.9)
                                        while True:
                                            wrongUser = 'on'
                                            try:
                                                chk_tele.wait_float(0.3,0.9)
                                                pg.click(fore.left+500,fore.top+300)
                                                refuseUserChkModal = driver.find_element(by=By.CSS_SELECTOR, value='.modal-title')
                                                refuseUserChkModal.click()
                                            except:
                                                break
                                except:
                                    pass
                                
                                try:
                                    manageText = driver.find_element(by=By.CSS_SELECTOR, value='.RightHeader .Transition__slide--active')
                                    if manageText.text == '관리자':
                                        break
                                except:
                                    pass
                                
                                try:
                                    menegerOkBtn = driver.find_element(by=By.CSS_SELECTOR, value='.Management .FloatingActionButton')
                                    if menegerOkBtn.get_attribute('title') == '저장':
                                        menegerOkBtn.click()
                                    chk_tele.wait_float(1.5,2.5)
                                except:
                                    pass
                                
                                
                                
                                
                            if wrongUser == 'on':
                                chk_tele.wrongUserWork(driver,fore,setUserName)
                                chkDbCount = 0
                                while True:
                                    chkDbCount += 1
                                    if dbSheet.cell(chkDbCount,4).value is not None:
                                        chkDb = re.sub(r'[^0-9]', '', str(dbSheet.cell(chkDbCount,4).value))
                                    if chkDb and chkDb in str(setUserName):
                                        break
                                dbSheet.cell(chkDbCount,8).value = 'V'
                                dbList.save('./db_list.xlsx')
                                continue
                                
                                
                            
                            
                            # 관리자 승격된 사람 클릭 관리자 권한 삭제 준비------------------------------
                            getSearchManager = ''
                            while True:
                                # print("관리자 승격된 사람 클릭 관리자 권한 삭제 준비")
                                chk_tele.wait_float(1.2,1.9)
                                try:
                                    delIcon = driver.find_element(by=By.CSS_SELECTOR, value='.destructive')
                                    if delIcon:
                                        break
                                except:
                                    pass
                                
                                try:
                                    for i in range(3):
                                        managerList = chk_tele.searchWaitElement('.Management .ListItem .fullName', driver)
                                        for manager in managerList:
                                            getManagerName = re.sub(r'[^0-9]', '', manager.text)
                                            if getManagerName == setUserName:
                                                getSearchManager = 'on'
                                                manager.click()
                                                break
                                    if getSearchManager == 'on':
                                        break
                                    else:
                                        chk_tele.wait_float('Telegram')
                                        pg.moveTo(fore.right - 150, fore.bottom - 300)
                                        pg.scroll(-1000)
                                                
                                except:
                                    pass
                                
                            # 관리자 권한 삭제 모달창 띄우기
                            chk_tele.searchAndClick('.Modal','.destructive',driver)
                            
                            # 모달창에서 관리자 권한 삭제 완료
                            chk_tele.searchTextAndClick('관리자', '.Modal .confirm-dialog-button.default.danger.text', driver)
                            
                        
                        # ★★★★★★★★★★★
                        else:
                            
                            if showLog:
                                print('그룹 참여 버튼 클릭')
                            chk_tele.searchAndClick('.AddChatMembers-inner .form-control', '#RightColumn .FloatingActionButton.revealed', driver)
                            
                            if showAlert:
                                pg.alert('그룹 참여 버튼 클릭 완료')
                            
                            findUser = ''
                            wrongUser = ''
                            while True:
                                if showLog:
                                    print("그룹 초대하기")
                                chk_tele.wait_float(0.9,1.2)
                                
                                # 초대 거절 설정해놓은 회원 거르기
                                try:
                                    refuseUserChkModal = driver.find_element(by=By.CSS_SELECTOR, value='.modal-title')
                                    if refuseUserChkModal and 'wrong' in refuseUserChkModal.text:
                                        chk_tele.wait_float(0.3,0.9)
                                        while True:
                                            wrongUser = 'on'
                                            try:
                                                chk_tele.wait_float(0.3,0.9)
                                                pg.click(fore.left+500,fore.top+300)
                                                chk_tele.wait_float(0.5,0.9)
                                                refuseUserChkModal = driver.find_element(by=By.CSS_SELECTOR, value='.modal-title')
                                                refuseUserChkModal.click()
                                            except:
                                                break
                                except:
                                    pass
                                
                                # 회원 없으면 걍 종료
                                try:
                                    nothingFound = driver.find_element(by=By.CSS_SELECTOR, value='.AddChatMembers-inner .no-results')
                                    if nothingFound:
                                        findUser = ''
                                        break
                                except:
                                    pass
                                
                                # 프로필 나오면 정상 패스~~
                                try:
                                    ProfilePhoto = driver.find_element(by=By.CSS_SELECTOR, value='.ProfilePhoto')
                                    if ProfilePhoto:
                                        break
                                except:
                                    pass
                                
                                
                                
                                try:
                                    chk_tele.wait_float(0.9,1.2)
                                    searchUserNameList = driver.find_elements(by=By.CSS_SELECTOR, value='.AddChatMembers-inner .ListItem')
                                    if len(searchUserNameList) == len(exceptUserList):
                                        findUser = ''
                                        break
                                    for userItem in searchUserNameList:
                                        userName = userItem.find_element(by=By.CSS_SELECTOR, value='.fullName')
                                        if userName.text[0:3] == '010' or userName.text[0:2] == '10':
                                            
                                            setUserName = re.sub(r'[^0-9]', '', userName.text)
                                            if setUserName in exceptUserList:
                                                continue
                                            userItem.click()
                                            findUser = 'on'
                                            break
                                except:
                                    pass
                                
                                
                                # 연락추가 버튼
                                try:
                                    chk_tele.wait_float(0.9,1.2)
                                    addUserBtn = driver.find_element(by=By.CSS_SELECTOR, value='#RightColumn .FloatingActionButton.revealed')
                                    addUserBtn.click()
                                except:
                                    pass
                            
                            if showAlert:
                                pg.alert(f'그룹 초대 완료 / 초대 거절 유저일 경우 : {wrongUser} // 유저를 찾음 여부 : {findUser} // ')
                            
                            if findUser == '':
                                if showLog:
                                    print('더이상 찾을 회원이 없음')
                                while True:
                                    if showLog:
                                        print("그룹 설정 메인 갈때까지 뒤로가기 클릭")
                                    chk_tele.wait_float(0.5,0.9)
                                    try:
                                        closeBtn = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.close-button')))
                                        getBtnText = closeBtn.get_attribute('title')
                                        if getBtnText == '닫기':
                                            memberList = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.content.members-list')))
                                            if memberList:
                                                break
                                        else:
                                            closeBtn.click()
                                    except:
                                        pass
                                if showAlert:
                                    pg.alert(f'우측 그룹 설정 메인으로 이동 완료')
                                    
                                while True:
                                    if showLog:
                                        print('연락처로 이동')
                                    try:
                                        chk_tele.wait_float(1.2,1.9)
                                        chatListWrap = driver.find_elements(by=By.CSS_SELECTOR, value='#LeftColumn-main > .Transition > div')
                                        chatList = chatListWrap[1].find_element(by=By.CSS_SELECTOR, value='.chat-list')
                                        chatList.click()
                                        break
                                    except:
                                        pass
                                    
                                    try:
                                        menuList = chk_tele.showTeleMenu(driver)
                                        menuList[2].click()
                                    except:
                                        chk_tele.wait_float(0.5,1.2)
                                        pg.click(fore.left+500,fore.top+300)
                                        chk_tele.wait_float(0.5,1.2)
                                
                                if showAlert:
                                    pg.alert(f'좌측 연락처로 이동 완료')
                                
                                while True:
                                    if showLog:
                                        print('연락처에 있는 유저 전부 삭제')
                                    try:
                                        chk_tele.wait_float(1.2,1.9)
                                        chatListWrap = driver.find_elements(by=By.CSS_SELECTOR, value='#LeftColumn-main > .Transition > div')
                                        chatList = chatListWrap[1].find_element(by=By.CSS_SELECTOR, value='.chat-list')
                                        if 'empty' in chatList.text:
                                            break
                                    except:
                                        pass
                                    
                                    try:
                                        chatListWrap = driver.find_elements(by=By.CSS_SELECTOR, value='#LeftColumn-main > .Transition > div')
                                        chatList = chatListWrap[1].find_elements(by=By.CSS_SELECTOR, value='.chat-list .ListItem')
                                        
                                        for chatRoom in chatList:
                                            nowUserName = chatRoom.text
                                            while True:
                                                chk_tele.wait_float(1.2,1.9)
                                                try:
                                                    getHeadName = driver.find_element(by=By.CSS_SELECTOR, value='.MiddleHeader .fullName')
                                                    if getHeadName.text in nowUserName:
                                                        break
                                                    
                                                except:
                                                    pass
                                                
                                                try:
                                                    chatRoom.click()
                                                except:
                                                    pass
                                            
                                            
                                            
                                            # 연락처 삭제 준비, 삭제 아이콘 나오게
                                            chk_tele.searchAndClick('.icon-delete', '.tools button', driver, 1, '.MiddleHeader .fullName')
                                            
                                            chk_tele.openModalAndDelete('회원 정보',driver)
                                            # # 연락처 삭제 모달창 띄우기
                                            # chk_tele.searchAndClick('.Modal', '.destructive', driver)
                                                
                                            # # 연락처 삭제 완료
                                            # chk_tele.searchTextAndClick('회원 정보', '.Modal .confirm-dialog-button.default.danger.text', driver)
                                        
                                    except:
                                        pass
                                
                                if showAlert:
                                    pg.alert('연락처 전부 삭제 완료')
                                
                                authSheet.cell(authCount, 6).value = f"{todayStr} 작업 완료"
                                authList.save('./auth_list.xlsx')
                                break
                            
                            if wrongUser == 'on':
                                chk_tele.wrongUserWork(driver,fore,setUserName)
                                chkDbCount = 0
                                while True:
                                    chkDbCount += 1
                                    if dbSheet.cell(chkDbCount,4).value is not None:
                                        chkDb = re.sub(r'[^0-9]', '', str(dbSheet.cell(chkDbCount,4).value))
                                    if chkDb and chkDb in str(setUserName):
                                        break
                                dbSheet.cell(chkDbCount,8).value = 'V'
                                dbList.save('./db_list.xlsx')
                                
                                if showAlert:
                                    pg.alert('초대 거절 유저 삭제 완료')
                                continue
                            
                        # while True:
                        #     if showLog:
                        #         print("그룹 설정 메인 갈때까지 뒤로가기 클릭")
                        #     chk_tele.wait_float(0.5,0.9)
                        #     try:
                        #         closeBtn = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.close-button')))
                        #         getBtnText = closeBtn.get_attribute('title')
                        #         if getBtnText == '닫기':
                        #             memberList = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.content.members-list')))
                        #             if memberList:
                        #                 break
                        #         else:
                        #             closeBtn.click()
                        #     except:
                        #         pass
                        
                        while True:
                            if showLog:
                                print("연락처 삭제 준비, 회원 클릭 (헤더에 회원 이라고 나타나게)")
                            
                            try:
                                chk_tele.wait_float(1.2,1.9)
                                userInfo = driver.find_element(by=By.CSS_SELECTOR, value='.RightHeader .Transition__slide--active')
                                if '회원' in userInfo.text:
                                    break
                            except:
                                pass
                            
                            try:
                                chk_tele.wait_float(1.2,1.9)
                                groupStatus = driver.find_element(by=By.CSS_SELECTOR, value='.group-status')
                                groupCount = re.sub(r'[^0-9]', '', groupStatus.text)
                                forCountTemp = int(groupCount) // 10
                                if forCountTemp < 2:
                                    forCount = 2
                                else:
                                    forCount = forCountTemp
                                
                                getSearchMember = ''
                                chk_tele.wait_float(0.9,1.5)
                                for i in range(forCount):
                                    
                                    # print('멤버 목록 찾아서 클릭 에러 구간')
                                    
                                    memberList = chk_tele.searchWaitElement('.content.members-list .ListItem', driver)
                                    for member in memberList:
                                        memberText = member.find_element(by=By.CSS_SELECTOR, value='.ChatInfo .Avatar').get_attribute('aria-label')
                                        if memberText is None:
                                            memberText = member.find_element(by=By.CSS_SELECTOR, value='.ChatInfo .Avatar img').get_attribute('alt')
                                            
                                        getmemberName = re.sub(r'[^0-9]', '', memberText)
                                        if getmemberName and getmemberName == setUserName:
                                            member.click()
                                            getSearchMember = 'on'
                                            break
                                    if getSearchMember == 'on':
                                        if showAlert:
                                            pg.alert('멤버를 찾았다.')
                                        break
                                    else:
                                        if showAlert:
                                            pg.alert('멤버 못찾음 스크롤 내려보자')
                                        pg.moveTo(fore.right - 150, fore.bottom - 300)
                                        pg.scroll(-1000)
                                        chk_tele.wait_float(1.2,1.9)
                                
                                if getSearchMember == '':
                                    chkDbCount = 0
                                    while True:
                                        chkDbCount += 1
                                        if dbSheet.cell(chkDbCount,4).value is not None:
                                            chkDb = re.sub(r'[^0-9]', '', str(dbSheet.cell(chkDbCount,4).value))
                                        if chkDb and chkDb in str(setUserName):
                                            break
                                    chk_tele.wait_float(0.3,0.9)
                                    if showLog:
                                        print('초대 불가 유저 체크')
                                    dbSheet.cell(chkDbCount,8).value = 'V'
                                    dbList.save('./db_list.xlsx')
                                    
                                    if showLog:
                                        print('초대 안되는 유저 예외 배열 추가')
                                    exceptUserList.append(setUserName)
                                    break
                            except Exception as e:
                                pg.alert(e)
                                pass
                        if getSearchMember == '':
                            authSheet.cell(authCount, 8).value = f"초대 기능 짤렸는지 확인"
                            authList.save('./auth_list.xlsx')
                            continue
                        # 여기가 난제!!!!!!!!!!!!! ★★★★★★★★★★★★★★★★★
                        # if getSearchMember == '':
                        #     authSheet.cell(authCount, 6).value = f"초대 기능 짤림"
                        #     authList.save('./auth_list.xlsx')
                        #     break
                        
                        # 연락처 삭제 준비, 삭제 아이콘 나오게
                        chk_tele.searchAndClick('.icon-delete', '.tools button', driver)
                        
                        chk_tele.openModalAndDelete('회원 정보',driver)
                        
                        
                        # 엑셀에 삭제된 연락처 진짜 이름 추가
                        reCount = 0
                        chkDb = ''
                        while True:
                            reCount += 1
                            chk_tele.wait_float(0.9,1.5)
                            if reCount > 3:
                                reCount = 0
                                chk_tele.wait_float('Telegram')
                                # print('새로고침!!!')
                                pg.press('F5')
                                chk_tele.wait_float(0.5,1.2)
                            
                            try:
                                # print('진짜 이름 구하기 에러체크')
                                getRealName = driver.find_element(by=By.CSS_SELECTOR, value='.MiddleHeader > .Transition > .Transition__slide--active > .chat-info-wrapper .fullName')
                                getReal = re.sub(r'[^\uAC00-\uD7A30-9a-zA-Z\s]', '', getRealName.text)
                                if getReal == setUserName or getReal == '':
                                    raise Exception('no match name')
                                else:
                                    chkDbCount = 0
                                    while True:
                                        chkDbCount += 1
                                        if dbSheet.cell(chkDbCount,4).value is not None:
                                            chkDb = re.sub(r'[^0-9]', '', str(dbSheet.cell(chkDbCount,4).value))
                                        if chkDb and chkDb in str(setUserName):
                                            break
                                    chk_tele.wait_float(0.3,0.9)
                                    dbSheet.cell(chkDbCount, 2).value = getReal
                                    dbList.save('./db_list.xlsx')
                                    break
                            except Exception as e:
                                if showLog:
                                    print(e)
                                # pg.alert('삭제중 에러발생! 현재 화면 캡쳐해서 관리자에게 문의 주세요!')
                                pass
                chk_tele.wait_float(0.5,1.2)
                driver.quit()
                chk_tele.wait_float(0.5,1.2)
                
        except Exception as e:
            print(e)
            if showAlert:
                if 'Windows' in e:
                    pg.alert('윈도우 관련 에서!! 이 에러 나면 그냥 패스하자!!!')
                    pg.alert(e)
                
            pg.alert('알수없는 에러발생! 현재 화면 캡쳐해서 관리자에게 문의 주세요!')
            
    sys.exit(0)
            
            
                
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
    
    
    
def authListChk():
    
    try:
        with open(f'./auth.txt', 'r', encoding='UTF8') as r:
            get_auth = re.sub(r'[/s]', '', r.read())
    except:
        with open(f'./auth.txt', 'r') as r:
            get_auth = re.sub(r'[/s]', '', r.read())
    
    get_mac = getmac.get_mac_address()
    
    webhook_url = "https://adpeak.kr/telework/gethook"
    data = {'get_auth' : get_auth, 'get_mac' : get_mac}
    requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)
    r = requests.post(webhook_url, data=json.dumps(data), headers={'Content-Type' : 'application/json'}, verify=False)
    wh_result = r.json()

    if wh_result['get_status'] == 'no':
        pg.alert('인증에 실패하였습니다. 관리자에게 문의해주세요')
        sys.exit(0)
    elif wh_result['get_status'] == 'retry':
        pg.alert('등록 되었습니다. 다시 시도 해주세요')
        sys.exit(0)
    elif wh_result['get_status'] == 'ok':
        func_url = wh_result['hidden_link']

    with httpimport.remote_repo(func_url):
        import chk_tele
        
    pcUser = getpass.getuser()
    authList = load_workbook('./auth_list.xlsx')
    authSheet = authList.active
    
    authCount = 1
    while True:
        authCount += 1
        profileNum = authSheet.cell(authCount,1).value
        if profileNum is None:
            pg.alert('완료되었습니다!!')
            break
        
        profileStatus = authSheet.cell(authCount,3).value
        if profileStatus is None or "X" in profileStatus:
            options = Options()
            user_data = f'C:\\Users\\{pcUser}\\AppData\\Local\\Google\\Chrome\\User Data\\default'
            service = Service(ChromeDriverManager().install())
            options.add_argument(f"user-data-dir={user_data}")
            options.add_argument(f'--profile-directory=Profile {profileNum}')
            driver = webdriver.Chrome(service=service, chrome_options=options)
            driver.set_window_size(1600, 800)
            driver.set_window_position(0,0)
            fore = pg.getActiveWindow()
            driver.get('https://web.telegram.org/z/')
            
            chk_tele.wait_float(1.9,2.5)
            chkAuth = pg.confirm(f'현재 접속한 아이디는 {profileNum} 입니다. 인증을 진행하시려면 yes를 / 확인만 하시려면 no를 클릭해주세요',buttons=['yes','no'])
            if chkAuth == 'yes':
                authSheet.cell(authCount,3).value = '인증완료'
                authList.save('./auth_list.xlsx')
                getAuthPhNum = pg.prompt(title='TITLE',default='',text=f'현재 접속한 아이디는 {profileNum} 입니다. 아래 칸에 국가번호 전화번호를 입력하시면 엑셀에 반영됩니다. (미작성시 패스~)')
                
                if getAuthPhNum == '' or getAuthPhNum is None:
                    pass
                else:
                    authSheet.cell(authCount,2).value = getAuthPhNum
                    authList.save('./auth_list.xlsx')
                    
            driver.quit()
    
    pg.alert('인증 작업이 완료 되었습니다. 엑셀 파일을 확인 해주세요!')
    
def delAuthList():
    pcUser = getpass.getuser()
    authList = load_workbook('./auth_list.xlsx')
    authSheet = authList['auth_id_list']
    
    authCount = 1
    while True:
        authCount += 1
        profileNum = authSheet.cell(authCount,1).value
        if profileNum is None:
            break
        profileStatus = authSheet.cell(authCount,3).value
        if profileStatus == '인증XX' or profileStatus is None:
            defTargetProfileFolder = f'C:\\Users\\{pcUser}\\AppData\\Local\\Google\\Chrome\\User Data\\default\\Profile {profileNum}'
            
            if os.path.isdir(defTargetProfileFolder):
                shutil.rmtree(defTargetProfileFolder)
            authSheet.cell(authCount,3).value = ''
            authList.save('./auth_list.xlsx')
            
    pg.alert('인증 목록 정리가 완료 되었습니다. 다시 인증을 진행 해주세요!')