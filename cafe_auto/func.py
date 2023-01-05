import random
import threading
import time
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
import sys
import os
from pathlib import Path
from typing import Optional
import requests
from bs4 import BeautifulSoup as bs
import json
import re
import pyautogui as pg
import pyperclip
import pywinauto
import pygetwindow as gw
import clipboard as cb
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import requests
import winsound as ws
import glob
import asyncio
import string




def cafe_join_btn(driver):
    searchElement('.btn_write', driver)
    try:
        joinBtn = driver.find_element(by=By.CSS_SELECTOR, value='.inner_box .btn_join')
    except:
        return
    
    joinBtn.click()
    
    
    searchElement('.cafe_info', driver)
    
    
    
    wait_float(0.5,1.2)
    
    
    while True:
        focus_window('네이버 카페')
        try:
            labelJoinNick = driver.find_element(by=By.CSS_SELECTOR, value='#label_join_nick').get_attribute('value')
            messageAlertText = driver.find_element(by=By.CSS_SELECTOR, value='.input_message').text
            
            if labelJoinNick is None or labelJoinNick == '' or '이미' in messageAlertText:
                ranVal = random.randrange(5,9)
                rand_str = ''
                for i in range(ranVal):
                    rand_str += str(random.choice(string.ascii_lowercase))
                wait_float(0.5,1.2)
                
                labelJoinNickArea = driver.find_element(by=By.CSS_SELECTOR, value='#label_join_nick')
                labelJoinNickArea.click()
                wait_float(0.5,1.2)
                focus_window('네이버 카페')
                wait_float(0.5,1.2)
                pg.hotkey('ctrl', 'a')
                wait_float(0.5,1.2)
                keyboard.write(text=rand_str, delay=0.3)
        except:
            pass
        

        
        for kk in range(0, 11):
            try:
                joinQuestion = driver.find_element(by=By.CSS_SELECTOR, value=f'#label_join_question_{kk}')
                if joinQuestion:
                    joinVal = joinQuestion.get_attribute('value')
                    if joinVal == '' or joinVal is None:
                        wait_float(0.5,1.2)
                        joinQuestion.click()
                        focus_window('네이버 카페')
                        wait_float(0.5,1.2)
                        keyboard.write(text='네 알겠습니다.', delay=0.3)
            except:
                pass
        
        
        for ii in range(0, 11):
            try:
                radioQuestion = driver.find_element(by=By.CSS_SELECTOR, value=f'#radio_join_question_{ii}_0')
                if radioQuestion:
                    wait_float(0.5,1.2)
                    radioQuestion.click()
                    wait_float(0.5,1.2)
            except:
                pass
        
        wait_float(2.5,3.2)
        
        
        
        
        try:
            joinFailBtn = driver.find_element(by=By.CSS_SELECTOR, value='.BaseButton—disabled')
            if joinFailBtn:
                continue
        except:
            joinSuccessBtn = driver.find_element(by=By.CSS_SELECTOR, value='.join_btn_box .ButtonBase--green')
            joinSuccessBtn.click()
            break
        
    searchElement('.section_cafe', driver)
    
    try:
        joinSuccessModal = driver.find_element(by=By.CSS_SELECTOR, value=f'.join_ly')
        if joinSuccessModal:
            modalClose = driver.find_element(by=By.CSS_SELECTOR, value=f'.btn_lyr_clse')
            modalClose.click()
    except:
        pass
    
            
    
        
        
    
    
            
            
            
            
    
    
    # driver.execute_script("arguments[0].scrollIntoView();", replySuccessBtn[0])
    
    
    
def login_step():
    
    chromeVersionChkPath = 'C:\\Users\\pcy\\AppData\\Local\\Google\\Chrome\\User Data\\default'
    # chromeVersionChkPath = 'C:\\Users\\드림모어\\AppData\\Local\\Google\\Chrome\\User Data\\Default'
    
    optimize_ex = load_workbook('./etc/naver_optimiz.xlsx')
    opt_ex = optimize_ex.active
    count = 0
    preIp = ''
    
    while True:
        count += 1
        while True:
            getIP = changeIp()
            if not preIp == getIP:
                preIp = getIP
                break
            
        uaSet = opt_ex.cell(count,1).value
        if uaSet is None:
            pg.alert('종료합니다~~')
            break
        
        options = Options()
        user_data = chromeVersionChkPath
        service = Service(ChromeDriverManager().install())
        options.add_argument(f"user-data-dir={user_data}")
        options.add_argument(f'--profile-directory={uaSet}')
        driver = webdriver.Chrome(service=service, chrome_options=options)
        driver.set_window_size(1180, 910)
        driver.set_window_position(0,0)
        
        driver.get('https://www.naver.com')
        
        nId = opt_ex.cell(count,2).value
        nPwd = opt_ex.cell(count,3).value
        naverLogin_pc(nId,nPwd,driver)
        
        pg.alert('대기요~~~ 할거 하고 엔터~~~~~')
        
        driver.quit()
        
    






def mobile_chrome(getDict):
    preIp = ''
    
    if getDict['ipval'] == 1:
        while True:
            getIP = changeIp()
            if not preIp == getIP:
                preIp = getIP
                break
    with open(f'./etc/useragent/useragent_all.txt') as f:
        ua_data = f.readlines()
        randomUaCount = random.randrange(0, len(ua_data))
        getUaData = ua_data[randomUaCount]

    options = Options()
    user_agent = getUaData
    options.add_argument('user-agent=' + user_agent)
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(chrome_options=options, service=service)
    driver.get('https://www.naver.com')
    pg.alert('할거 하고 끝내기~')
    
    
    
def cafe_write_mobile(nBoardName,chk_extesion,driver):
    
    
    print('카페 진입 완료')
    
    try:
        joinBtn = driver.find_element(by=By.CSS_SELECTOR, value='.inner_box .btn_join')
        if joinBtn:
            return
    except:
        pass

    writeBtn = searchElement('.inner_box .btn_write', driver)
    untilEleGone(writeBtn[0], '.inner_box .btn_write', driver)

    # 메뉴 선택하기
    selectBox = searchElement('.selectbox', driver)
    untilEleShow(selectBox[0], '.layer_dimmed', driver)
    selBoard = searchElement('.select_board li', driver)
    for board in selBoard:
        if board.text == nBoardName:
            board.click()
            break

    print('메뉴 선택 완료')

    # 메뉴 선택 완료! 글쓰기 시작~~~~
    with open("./etc/content/write_content.txt", "r") as f:
        getContents = f.readlines()

    chkImg = [0, 0, 0, 0, 0, 0]

    # 글 작성 전 창 체크
    focus_window('글쓰기')

    articleWriteFormSubject = searchElement('.ArticleWriteFormSubject', driver)
    articleWriteFormSubject[0].click()
    keyboard.write(text=getContents[0], delay=0.3)

    oneEditor = searchElement('#one-editor', driver)
    oneEditor[0].click()
    for i, line in enumerate(getContents):
        if i == 0:
            continue
        try:
            line_temp = line.replace('\n', '')
            chkImg = line_temp.split('|')
        except:
            pass

        if chkImg[0] == 'img_line':
            imageUpload = searchElement('.se-toolbar-item-image', driver)
            imageUpload[0].click()

            nowPath = os.getcwd()

            imagePath = nowPath + "\etc\content\images"
            imageList = os.listdir(imagePath)

            while True:
                getImage = imageList[random.randrange(
                    0, len(imageList))]
                getImage_ex = getImage.split('.')[-1]
                if getImage_ex in chk_extesion:
                    break
            wait_float(1.5, 2.2)
            pyperclip.copy(imagePath)
            wait_float(0.5, 0.9)
            pg.hotkey('ctrl', 'v')
            wait_float(0.5, 0.9)
            pg.press('enter')

            wait_float(0.9, 1.6)
            pyperclip.copy(getImage)
            wait_float(0.5, 0.9)
            pg.hotkey('ctrl', 'v')
            wait_float(0.5, 0.9)
            pg.press('enter')

            wait_float(3.5, 4.8)
            # 끝난다음 초기화
            chkImg = [0, 0, 0, 0, 0, 0]
        elif line == 'enter':
            pg.press('enter')
        else:
            keyboard.write(text=line, delay=0.05)
        wait_float(0.7, 1.3)
    successBtn = searchElement('.GnbBntRight__green', driver)
    untilEleGone(successBtn[0], '.GnbBntRight__green', driver)

    print('글쓰기 완료')

    getLinkMore = searchElement('.btn_aside .more', driver)
    getLinkMore[0].click()

    getLink = searchElement('.layer_list li', driver)
    getLink[3].click()

    getLinkData = cb.paste()
    
    with open('./etc/work_link.txt', 'a') as f:
        f.write('\n')
        f.write(getLinkData)

    with open('./etc/work_link.txt', 'r') as f:
        chkLines = f.readlines()
        
    delLine = ''
    if len(chkLines) > 6:
        getDelOptimizeReplyNum_temp = chkLines[0].split('/')
        getDelOptimizeReplyNum = getDelOptimizeReplyNum_temp[-1]
        if os.path.exists(f'./etc/content/temp_reply/{getDelOptimizeReplyNum}.txt'):
            os.remove(
                f'./etc/content/temp_reply/{getDelOptimizeReplyNum}.txt')
        delLine = chkLines[0]
        del chkLines[0]

    with open('./etc/work_link.txt', 'w') as f:
        f.writelines(''.join(chkLines))
    goToHome = searchElement('.header h1', driver)
    untilEleGone(goToHome[0], '.post_title', driver)
    
    return [getLinkData,delLine]
 

def cafe_re_reply_mobile(driver,cafeName):
    
    driver.get(cafeName)
    
    while True:
        wait_float(1.7,2.9)
        gnb_ham = driver.find_element(by=By.CSS_SELECTOR, value='.gnb_ham')
        gnb_ham.click()
        
        prevent_scroll = driver.find_element(by=By.CSS_SELECTOR, value='.prevent_scroll')
        if prevent_scroll:
            break
    
    wait_float(0.7,1.8)
    profileBtn = searchElement('.login_after .thmb', driver)
    profileBtn[0].click()
    
    wait_float(0.7,1.8)
    firstList = searchElement('.list_area li', driver)
    firstList[0].click()
    
    
    replyBtn = searchElement('.f_reply', driver)
    replyBtn[0].click()
    
    
    commentList = searchElement('.comment_list li', driver)
    
    replyNum = 0
    for comment in commentList:
        chkClassList = comment.get_attribute('class')
        if 'mine' in chkClassList:
            continue
        replyNum += 1
        
    if replyNum < 1:
        return

    eleCount = 0
    for i in range(replyNum):
        comment_list = driver.find_elements(by=By.CSS_SELECTOR, value=".comment_list li")
        while True:
            chkClassList = comment_list[eleCount].get_attribute('class')
            if 'mine' in chkClassList or 'reply' in chkClassList:
                eleCount += 1
                continue
            else:
                break
        
        while True:
            wait_float(0.7,1.8)
            btn_write = comment_list[eleCount].find_element(by=By.CSS_SELECTOR, value=".btn_write")
            btn_write.click()
            
            try:
                writeBox = driver.find_element(by=By.CSS_SELECTOR, value=".type__open")
                if writeBox:
                    break
            except:
                continue
        
        while True:
            wait_float(0.7,1.8)
            stickerBtn = writeBox.find_element(by=By.CSS_SELECTOR, value=".TownCommentWriteAttachSticker")
            stickerBtn.click()
            
            sticker_list = writeBox.find_elements(by=By.CSS_SELECTOR, value=".sticker_list li")
            if len(sticker_list) > 0:
                break
        
        wait_float(0.7,1.8)
        ranVal = random.randrange(0,len(sticker_list))
        sticker_list[ranVal].click()
        
        wait_float(0.7,1.8)
        btn_done = driver.find_element(by=By.CSS_SELECTOR, value=".btn_done")
        btn_done.click()
        
        wait_float(1.7,2.8)
        eleCount += 1
    

def cafe_reply_mobile(driver):
    
    
    
    with open(f'./etc/work_link.txt') as f:
        workLinkList = f.readlines()
        
    for i, workLink in enumerate(workLinkList):
        exceptVal = ''
        workLink_temp = workLink.replace('\n', '')
        workLinkOn = workLink_temp.split('/')[-1]
        boardListAll = searchElement('.list_area li', driver)
        for boardList in boardListAll:
            try:
                chkBoardLink = boardList.find_element(
                by=By.CSS_SELECTOR, value='.txt_area').get_attribute('href')
            except:
                exceptVal = 'on'
                break
            chkVal = workLinkOn in chkBoardLink

            if chkVal:
                try:
                    untilEleGone(boardList, '.txt_area', driver)
                except:
                    exceptVal = 'on'
                    break
                pg.moveTo(300, 500)
                randomFor = random.randrange(2, 5)
                for i in range(1, randomFor):
                    wait_float(1.5, 2.5)
                    pg.scroll(-500)
                break
        
        try:
            error_page = driver.find_element(by=By.CSS_SELECTOR, value='.EmptyMessageBox')
            if error_page:
                preBtn = driver.find_element(by=By.CSS_SELECTOR, value='.ButtonBase--gray')
                preBtn.click()
        except:
            pass
            
        # 게시글 클릭 완료 댓글 쓰기 시작!
        if exceptVal == 'on':
            continue
        randomActionVal = random.randrange(1, 4)
        print(randomActionVal)
        if randomActionVal == 1:
            
            # 댓글 작성 전 창 체크
            replyGoBtn = searchElement('.f_reply', driver)
            untilEleShow(replyGoBtn[0], '.HeaderIcon', driver)

            wait_float(0.5, 0.9)
            replyBtn = searchElement('.comment_textarea', driver)
            replyBtn[0].click()

            getReply = ""
            print(workLinkOn)
            if os.path.exists(f'./etc/content/temp_reply/{workLinkOn}.txt'):

                with open(f'./etc/content/temp_reply/{workLinkOn}.txt', 'r') as f:
                    getTempReplys = f.readlines()
                    getTempNum = getTempReplys[0].replace('\n', '')
                    try:
                        getReply = getTempReplys[int(getTempNum) + 1]
                        getTempReplys[0] = str(int(getTempNum) + 1) + '\n'
                    except:
                        getReply = ""
                wait_float(0.2, 0.7)
                if getReply != "":
                    with open(f'./etc/content/temp_reply/{workLinkOn}.txt', 'w') as f:
                        f.writelines(''.join(getTempReplys))
            wait_float(0.2, 0.7)
            if getReply == "":
                with open(f'./etc/all_reply.txt', 'r') as f:
                    getTempReplysAll = f.readlines()
                    getTempRanNum = random.randrange(
                        0, len(getTempReplysAll))
                    getReply = getTempReplysAll[getTempRanNum]

            focus_window('카페')
            keyboard.write(text=getReply, delay=0.1)
            replySuccessBtn = searchElement('.btn_done', driver)
            untilEleGone(replySuccessBtn[0], '.btn_done', driver)

            goToPostiongBtn = searchElement('.HeaderGnbLeft', driver)
            untilEleGone(goToPostiongBtn[0], '.HeaderGnbLeft', driver)

        goToHome = searchElement('.header h1', driver)
        untilEleGone(goToHome[0], '.post_title', driver)



def cafe_re_reply_pc(cafeAllInfo,driver):
    navItem = searchElement('.nav_item', driver)
    for mitem in navItem:
        if mitem.text == '카페':
            mitem.click()
            break

    cafeList = searchElement('.user_mycafe_info', driver)
    getInCafe = ""
    for cafeOn in cafeList:
        if cafeAllInfo[0] in cafeOn.text:
            cafeOn.click()
            getInCafe = "on"
            break

    if getInCafe == "":
        driver.get(cafeAllInfo[1])
    else:
        driver.switch_to.window(driver.window_handles[1])
        driver.close()
        driver.switch_to.window(driver.window_handles[0])

    searchElement('.cafe-write-btn', driver)
    
    myActionTab = searchElement('#cafe-info-data .tit-action', driver)
    myActionTab[0].click()
    myPostList = searchElement('#member-action-data .info2 a', driver)
    myPostList[0].click()
    
    
    refreshCount = 0
    while True:
        refreshCount += 1
        if refreshCount % 5 == 0:
            pg.press('F5')
        print('씨발좆 111')
        wait_float(0.5,0.9)
        driver.switch_to.default_content()
        try:
            driver.switch_to.frame('cafe_main')
        except:
            continue
        
        try:
            chkProfile = driver.find_element(by=By.CSS_SELECTOR, value=".MemberProfile")
            if chkProfile:
                break
        except:
            continue
    
    wait_float(0.2,0.7)
    articleBoard = driver.find_element(by=By.CSS_SELECTOR, value=".article-board")
    wait_float(0.2,0.7)
    article = articleBoard.find_element(by=By.CSS_SELECTOR, value=".td_article .article")
    wait_float(0.2,0.7)
    article.click()
    
    
    refreshCount = 0
    while True:
        refreshCount += 1
        if refreshCount % 5 == 0:
            pg.press('F5')
        print('씨발좆 222')
        wait_float(0.5,0.9)
        
        try:
            driver.switch_to.window(driver.window_handles[1])
            driver.switch_to.default_content()
            driver.switch_to.frame('cafe_main')
            wait_float(0.5,0.9)
            articleContentBox = driver.find_element(by=By.CSS_SELECTOR, value=".ArticleContentBox")
            if articleContentBox:
                break
        except:
            continue
        
        
    
    commentItem = driver.find_elements(by=By.CSS_SELECTOR, value=".CommentItem")
    
    replyNum = 0
    mineNum = 0
    for val in commentItem:
        chkClassList = val.get_attribute('class')
        if 'mine' in chkClassList:
            mineNum += 1
            continue
        replyNum += 1
    
    if replyNum < 2 or mineNum > 3:
        wait_float(1.3,2.5)
        driver.switch_to.window(driver.window_handles[1])
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        return
    
    eleCount = 0
    for i in range(replyNum):
        
        commentItem = searchElement('.CommentItem', driver)
        while True:
            chkClassList = commentItem[eleCount].get_attribute('class')
            if 'mine' in chkClassList or 'reply' in chkClassList:
                eleCount += 1
                continue
            else:
                break
        wait_float(0.5,1.3)
        
        refreshCount = 0
        while True:
            refreshCount += 1
            if refreshCount % 5 == 0:
                pg.press('F5')
            wait_float(0.5,1.2)
            print('씨발좆 333')
            focus_window('카페')
            replyBtn = commentItem[eleCount].find_element(by=By.CSS_SELECTOR, value=".comment_info_button")
            replyBtn.click()
            
            writeBox = driver.find_element(by=By.CSS_SELECTOR, value=".CommentItem--reply .CommentWriter")
            if writeBox:
                writeBox.click()
                break
        
        wait_float(0.5,1.2)
        
        while True:
            stickerBtn = writeBox.find_element(by=By.CSS_SELECTOR, value=".button_sticker")
            wait_float(0.5,1.2)
            stickerBtn.click()
            try:
                stickerBox = writeBox.find_element(by=By.CSS_SELECTOR, value=".CommentLineSticker")
                if stickerBox:
                    break
            except:
                continue
        
        
        while True:
            stickerList = writeBox.find_elements(by=By.CSS_SELECTOR, value=".se2_linesticker_list li")
            wait_float(0.5,1.2)
            if len(stickerList) < 3:
                continue
            getRan = random.randrange(0,len(stickerList))
            stickerList[getRan].click()
            
            try:
                replySuccessBtn = writeBox.find_element(by=By.CSS_SELECTOR, value=".btn_register.is_active")
                wait_float(0.5,1.2)
                if replySuccessBtn:
                    replySuccessBtn.click()
                    break
            except:
                continue
        wait_float(1.5,2.3)
        eleCount += 1
    
    while True:
        wait_float(0.5,1.3)
        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[1])
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
        else:
            break
    
    

def cafe_write_pc(cafeAllInfo,writeCount,driver):
    
    dirList = os.listdir(f"{os.getcwd()}\\etc\\content\\opt\\id_{writeCount}")
    
    for dir in dirList:

        with open(f'./etc/content/opt/id_{writeCount}/{dir}/content.txt', 'r') as f:
            getContents = f.readlines()
            
        cafeInfo = getContents[0].split('|')
        driver.get(cafeInfo[0])
        
        workBoardLink = searchElement(f'#menuLink{cafeInfo[2]}', driver)
        workBoardLink[0].click()
        
        while True:
            try:
                driver.switch_to.frame('cafe_main')
                break
            except:
                continue

        # 카페에 글 작성하기

        workBoardWriteBtn = searchElement('#writeFormBtn', driver)
        workBoardWriteBtn[0].click()

        wait_float(1.5,2.5)
        driver.switch_to.window(driver.window_handles[1])

        

        subjectArea = searchElement('.FlexableTextArea', driver)
        subjectArea[0].click()
        keyboard.write(text=getContents[1], delay=0.05)
        wait_float(0.8,1.9)

        textArea = searchElement('.se-content', driver)
        textArea[0].click()
                
        for i, getline in enumerate(getContents):
            if i < 2:
                continue
            getline = getline.replace('\n', '')
            getImgAction = getline.split('|')
            if getline == 'enter':
                pg.press('enter')
                wait_float(0.5, 0.9)
            elif getImgAction[0] == 'img_line':
                imgBtn = searchElement('.se-document-toolbar li', driver)
                imgBtn[0].click()
                nowPath = os.getcwd()
                imagePath = nowPath + f"\etc\content\opt\id_{writeCount}\{dir}"
                wait_float(1.5, 2.2)
                pyperclip.copy(imagePath)
                wait_float(0.5, 0.9)
                pg.hotkey('ctrl', 'v')
                wait_float(0.5, 0.9)
                pg.press('enter')
                wait_float(0.9, 1.6)
                pyperclip.copy(getImgAction[1])
                wait_float(0.5, 0.9)
                pg.hotkey('ctrl', 'v')
                wait_float(0.5, 0.9)
                pg.press('enter')
                wait_float(3.5, 4.8)
                
                
                while True:
                    wait_float(1.2,2.3)
                    imgLength = driver.find_elements(by=By.CSS_SELECTOR, value=".se-drop-indicator img")
                    if len(imgLength) < 1:
                        continue
                    else:
                        break
            else:
                keyboard.write(text=getline, delay=0.05)
                wait_float(0.5, 0.9)
                pg.press('enter')
                
                
        BaseButton = searchElement('.BaseButton', driver)
        BaseButton[0].click()
        wait_float(2.5,3.3)
        
        cafe_id = cafeAllInfo[1].split('/')[-1]
        
        while True:
            try:
                driver.switch_to.frame('cafe_main')
                break
            except:
                continue
            
        BaseButton = searchElement('.button_url', driver)
        
        
        
        if cafe_id in getContents[0]:
            
            while True:
                BaseButton = searchElement('.button_url', driver)
                
                wait_float(1.5,2.5)
                BaseButton[0].click()
                wait_float(0.5,0.9)
                getLinkData = cb.paste()
                wait_float(0.5,0.9)
                
                wait_float(0.5,0.9)
                with open('./etc/work_link.txt', 'r') as f:
                    chkLines = f.readlines()
                
                if len(chkLines) == 0:
                    with open('./etc/work_link.txt', 'a') as f:
                        f.write(getLinkData)
                else:
                    with open('./etc/work_link.txt', 'a') as f:
                        f.write('\n')
                        f.write(getLinkData)
                wait_float(0.5,0.9)
                
                with open('./etc/work_link.txt', 'r') as f:
                    chkLinks = f.read()
                    
                wait_float(0.5,0.9)
                
                if getLinkData in chkLinks:
                    pg.alert('들어있냐?!?!?!?!?!?')
                    break
                else:
                    continue
            
            if os.path.isfile(f'./etc/content/opt/id_{writeCount}/{dir}/reply.txt') and cafe_id in getContents[0]:
                with open(f'./etc/content/opt/id_{writeCount}/{dir}/reply.txt', 'r') as f:
                    getTempReplys = f.readlines()
                    getTempReplys.insert(0, '0\n')
                    getTempReplysName_temp = getLinkData.split('/')
                    getTempReplysName = getTempReplysName_temp[-1]
                    
                with open(f'./etc/content/temp_reply/{getTempReplysName}.txt', 'w') as f:
                    f.writelines(''.join(getTempReplys))
            
            
            with open('./etc/work_link.txt', 'r') as f:
                linkLines = f.readlines()
                
            wait_float(0.5,1.2)
            
            if len(linkLines) > 6:
                setLines = linkLines[-6:]
                
                linkContent = ''
                for linkline in setLines:
                    linkContent = linkContent + linkline
                with open('./etc/work_link.txt', 'w') as w:
                    w.write(linkContent)
            # worklink 설정 끝~~~~~~~~~~~
                
            wait_float(0.5,0.9)
            driver.close()
            
            while True:
                try:
                    driver.switch_to.window(driver.window_handles[0])
                    driver.switch_to.frame('cafe_main')
                    break
                except:
                    continue

            # 글쓰기 및 worklink에 링크 추가 완료 PC버전 댓글 달기 GO!
            with open(f'./etc/work_link.txt') as f:
                workLinkList = f.readlines()
                
            for i, workLink in enumerate(workLinkList):
                driver.switch_to.default_content()
                wait_float(0.3,0.9)
                workBoardLink = searchElement('#menuLink0', driver)
                workBoardLink[0].click()
                
                workLink_temp = workLink.replace('\n', '')
                workLinkOn = workLink_temp.split('/')[-1]
                
                while True:
                    try:
                        driver.switch_to.frame('cafe_main')
                        break
                    except:
                        continue
                articleDiff = searchElement('.article-board', driver)
                articleList = articleDiff[1].find_elements(by=By.CSS_SELECTOR, value=".td_article")
                for uu, article in enumerate(articleList):
                    getArticleHref = article.find_element(by=By.CSS_SELECTOR, value=".article").get_attribute('href')
                    getArticleHref = getArticleHref.split('id=')[-1].split('&')[0]
                    if workLinkOn == getArticleHref:
                        
                        while True:
                            tempArticleWrap = searchElement('.article-board', driver)
                            articleVal = tempArticleWrap[1].find_elements(by=By.CSS_SELECTOR, value=".td_article .article")
                            articleVal[uu].click()
                            wait_float(1.5,2.2)
                            try:
                                driver.find_element(by=By.CSS_SELECTOR, value=".article_header")
                                break
                            except:
                                continue
                            
                        focus_window('카페')    
                        pg.moveTo(200, 500)
                        
                        randomFor = random.randrange(3, 6)
                        for i in range(1, randomFor):
                            wait_float(1.5, 2.5)
                            pg.scroll(-500)
                        
                        randomActionVal = random.randrange(1, 4)
                        if randomActionVal == 1:
                            getReply = ""
                            if os.path.exists(f'./etc/content/temp_reply/{workLinkOn}.txt'):
                                with open(f'./etc/content/temp_reply/{workLinkOn}.txt', 'r') as f:
                                    getTempReplys = f.readlines()
                                    getTempNum = getTempReplys[0].replace('\n', '')
                                    try:
                                        getReply = getTempReplys[int(getTempNum) + 1]
                                        getTempReplys[0] = str(int(getTempNum) + 1) + '\n'
                                    except:
                                        getReply = ""
                                wait_float(0.2, 0.7)
                                if getReply != "":
                                    with open(f'./etc/content/temp_reply/{workLinkOn}.txt', 'w') as f:
                                        f.writelines(''.join(getTempReplys))
                            wait_float(0.2, 0.7)
                            if getReply == "":
                                with open(f'./etc/all_reply.txt', 'r') as f:
                                    getTempReplysAll = f.readlines()
                                    getTempRanNum = random.randrange(
                                        0, len(getTempReplysAll))
                                    getReply = getTempReplysAll[getTempRanNum]
                                    
                            wait_float(1.5,2.8)
                            try:
                                replyArea = driver.find_element(by=By.CSS_SELECTOR, value=".comment_inbox_text")
                                replyArea.click()
                                keyboard.write(text=getReply, delay=0.03)
                                wait_float(1.5,2.5)
                                replySuccessBtn = searchElement('.btn_register', driver)
                                driver.execute_script("arguments[0].scrollIntoView();", replySuccessBtn[0])
                                replySuccessBtn[0].click()
                                wait_float(2.5,3.5)
                                print(replyArea)
                            except:
                                pass
                        break
    

def simple_writer():
    with open('./write_content.txt', 'r') as r:
        all_content = r.readlines()
        
    for i,line in enumerate(all_content):
        line = line.replace('\n','')
        if i == 0:
            pg.alert('엔터 친 후 제목 부분을 클릭해주세요!')
            keyboard.write(text=line, delay=0.1)
            pg.alert('제목 작성 완료! 엔터 친 후 본문 부분을 클릭해주세요!')
        elif 'image' in line:
            pg.alert('이미지를 첨부해주세요. 첨부 후 엔터 치고 본문 부분 클릭해주세요!')
        elif 'enter' in line:
            pg.press('enter')
        else:
            keyboard.write(text=line, delay=0.1)
            wait_float(0.7,1.3)
            pg.press('enter')
            
            
            
        
    


# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>함수 시작염

# 상품 들어가서 스크롤 내리고 나오기


# async def wait헤arr, ex, i, chk):
#     target_click = int(ex.cell(i, 4).value)
#     now_click = ex.cell(i, 5).value

#     if now_click is None:
#         ex.cell(i, 5).value = 0
#         now_click = 0
#     now_click = int(now_click)
#     if chk == 'Y':
#         if now_click < target_click:
#             arr.append(i)
#     else:
#         if now_click >= target_click:
#             arr.append(i)


# async def playAsync_getArr(arr, ex, linkCount, chk):
#     try:
#         await asyncio.gather(*[waitPrint(arr, ex, i, chk) for i in range(1, linkCount + 1)])
#     except:
#         pass


# 결과 값(workarr) 을 가지고 해당 인덱스의 엑셀에 1씩 더하기
# async def linkExcelPlus(ex, val):
#     setVal = ex.cell(val, 5).value
#     ex.cell(val, 5).value = setVal + 1


# async def playAsync_plusArr(arr, ex):
#     try:
#         await asyncio.gather(*[linkExcelPlus(ex, val) for val in arr])
#     except:
#         pass


def naverLogin_mobile(load_id, load_pass, driver):
    search_bar = searchElement(".sch_ico_aside",driver)
    search_bar[0].click()
    login_btn = searchElement(".ss_profile_wrap",driver)
    login_btn[0].click()

    # 로그인 부분

    focus_window('로그인')
    while True:
        searchElement("#id",driver)

        pyperclip.copy(load_id)
        id_input = driver.find_elements(by=By.CSS_SELECTOR, value="#id")
        id_input[0].click()
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'a')
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'v')
        wait_float(0.4, 0.7)

        pyperclip.copy(load_pass)
        pw_input = driver.find_elements(by=By.CSS_SELECTOR, value="#pw")
        pw_input[0].click()
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'a')
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'v')
        wait_float(0.4, 0.7)
        id_input_value = id_input[0].get_attribute('value')
        if id_input_value:
            pg.hotkey('enter')
            wait_float(0.5, 1.0)
        else:
            continue

        asideChk = 0
        noProblem = ""
        passExit = ""
        while 2 > asideChk:
            try:
                waitAside = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".aside_wrap")))
                if waitAside is not None:
                    noProblem = "on"
                    break
            except:
                asideChk += 1

        if noProblem != "on":
            searchElement("#header",driver)

        try:
            newDevice = driver.find_elements(
                by=By.CSS_SELECTOR, value=".btn_white")
            newDevice[0].click()
        except:
            pass

        try:
            greenBtn = driver.find_elements(
                by=By.CSS_SELECTOR, value=".btn_next")
            greenBtn[0].click()
            passExit = "on"
        except:
            pass

        try:
            protectId = driver.find_elements(
                by=By.CSS_SELECTOR, value=".ico_warning2")
            if protectId:
                return "보호조치"
        except:
            pass

        try:
            sleepId = driver.find_elements(
                by=By.CSS_SELECTOR, value=".warning_v2")
            if sleepId:
                return "휴면아이디"
        except:
            pass

        try:
            unPwd = driver.find_elements(
                by=By.CSS_SELECTOR, value=".error_message")
            if unPwd:
                return "비번틀림"
            # 다시 로그인 어쩌구......
        except:
            pass
        try:
            unPwd = driver.find_elements(
                by=By.CSS_SELECTOR, value=".action_inner")
            if unPwd:
                return "비정상적 활동"
            # 다시 로그인 어쩌구......
        except:
            pass

        if passExit != "on":
            goToMain = searchElement(".ah_close",driver)
            goToMain[0].click()
            wait_float(0.4, 0.7)

        time.sleep(5)
        break
        # 로그인 부분 끝



def naverLogin_pc(nId,nPwd,driver):
    # 네이버 로그인~~~~~~
    loginBtn = searchElement('.sc_login', driver)
    loginBtn[0].click()
    
    searchElement('#id', driver)
    focus_window('로그인')
    wait_float(0.3,0.9)
    while True:
        pg.click(50,400)
        inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
        inputId.click()
        wait_float(0.3,0.9)
        cb.copy(nId)
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'a')
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'v')
        inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
        if inputId.get_attribute('value') != "":
            break
    while True:
        inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
        inputPw.click()
        wait_float(0.3,0.9)
        cb.copy(nPwd)
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'a')
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'v')
        inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
        if inputPw.get_attribute('value') != "":
            break
    
    btnLogin = searchElement('.btn_login', driver)
    btnLogin[0].click()
    
    
    while True:
        try:
            WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#query")))
            break
        except:
            driver.get('https://www.naver.com')

def changeIp():
    os.system('adb server start')
    client = AdbClient(host="127.0.0.1", port=5037)
    device = client.devices()  # 디바이스 1개
    ondevice = device[0]
    ondevice.shell("input keyevent KEYCODE_POWER")
    ondevice.shell("svc data disable")
    ondevice.shell("settings put global airplane_mode_on 1")
    ondevice.shell(
        "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

    ondevice.shell("svc data enable")
    ondevice.shell("settings put global airplane_mode_on 0")
    ondevice.shell(
        "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
    time.sleep(3)
    
    while True:
        try:
            wait_float(0.5, 0.9)
            getIp = requests.get("https://api.ip.pe.kr/json/").json()['ip']
            if getIp is not None:
                break
        except:
            continue
    print(getIp)
    return getIp



def changeIpSpeed():
    
    
    os.system('adb server start')
    client = AdbClient(host="127.0.0.1", port=5037)
    device = client.devices()  # 디바이스 1개
    ondevice = device[0]
    while True:
        try:

            ondevice.shell("input keyevent KEYCODE_POWER")
            ondevice.shell("svc data disable")
            ondevice.shell("settings put global airplane_mode_on 1")
            ondevice.shell(
                "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

            ondevice.shell("svc data enable")
            ondevice.shell("settings put global airplane_mode_on 0")
            ondevice.shell(
                "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
            time.sleep(3)
            while True:
                try:
                    wait_float(0.5, 0.9)
                    getIp = requests.get("https://api.ip.pe.kr/json/").json()['ip']
                    if getIp is not None:
                        break
                except:
                    continue
        except:

            while True:
                try:
                    wait_float(0.5, 0.9)
                    getIp = requests.get("https://api.ip.pe.kr/json/").json()['ip']
                    if getIp is not None:
                        break
                except:
                    continue
                
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service)
        driver.get('https://fast.com/ko/')
        searchElement('.speed-results-container',driver)
        time.sleep(5)
        getInternetRapidEle = searchElement('.speed-results-container',driver)
        getInternetRapid = getInternetRapidEle[0].text
        if float(getInternetRapid) < 2.7 or float(getInternetRapid) > 100.0:
            driver.quit()
            continue
        else:
            driver.quit()
            break

    return getIp

def searchElement(ele,driver):
    wait_float(0.3, 0.7)
    re_count = 0
    element = ""
    while True:
        re_count += 1
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
            pg.press('F5')
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            pass

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.3, 0.7)
    return selected_element

def searchElement_seterr(ele,driver):
    wait_float(0.3, 0.7)
    re_count = 0
    element = ""
    while True:
        if re_count > 10:
            return 'error'
        re_count += 1
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
            pg.press('F5')
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            pass

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.3, 0.7)
    return selected_element


# def untilEleShow(clickEle, searchEle):
#     while True:
#         try:
#             clickEle.click()
#             time.sleep(1)
#             try:
#                 btnEle = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
#                 if btnEle is not None:
#                     return
#             except:
#                 continue
#         except:
#             pass

def untilEleShow(clickEle, searchEle, driver):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        try:
            btnEle = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is not None:
                return
        except:
            continue


# def untilEleGone(clickEle, searchEle):
#     while True:
#         try:
#             clickEle.click()
#             time.sleep(1)
#             try:
#                 btnEle = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
#                 if btnEle is None:
#                     return
#             except:
#                 return
#         except:
#             pass

def untilEleGone(clickEle, searchEle, driver):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass

        try:
            btnEle = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is None:
                return
        except:
            return


def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)


def exitApp():
    pg.alert(text='프로그램을 종료합니다.', title='제목입니다.', button='OK')
    sys.exit(0)

def focus_window(winName):
    while True:
        activeName = str(pg.getActiveWindow())
        print('에러 11111')
        wait_float(0.3,0.9)
        if winName in activeName:
            return
        
        print('에러 22222')
        winList = pg.getAllWindows()
        wait_float(0.3,0.9)
        print('에러 33333')
        for win in winList:
            print(win.title)
            if winName in win.title:
                pywinauto.application.Application().connect(handle=win._hWnd).top_window().set_focus()
                win.activate()
                wait_float(0.7,1.3)
                break


BASE_DIR = Path(__file__).resolve().parent


async def getEmptyArr(setNum, exName):
    getArr = []
    asyncio.gather(*[busyFunc(i, getArr, exName) for i in range(1, setNum)])
    return getArr


async def busyFunc(i, getArr, exName):
    getTime = exName.cell(i, 4).value

    try:
        if getTime is None:
            getTime = datetime.now().date()
        if isinstance(getTime, datetime):
            getTime = getTime.date()
        compareTime = datetime.now() - timedelta(days=3)
        if exName.cell(i, 4).value is None or getTime <= compareTime.date():
            getArr.append(i)
    except:
        pass


def getExLength(exName):
    ExLength = 0
    while True:
        ExLength += 1
        if exName.cell(ExLength, 2).value is None:
            break
    return ExLength


def getUaNum():
    with open("./etc/useragent/useragent_all.txt", "r") as f:
        fArr = f.readlines()
        fCount = len(fArr)
        uaSet = random.randrange(0, fCount)
    return uaSet


# def mainToCafe(driver):
#     shs_item = searchElement('.shs_item',driver)
#     for item in shs_item:
#         chkCafe = item.find_element(
#             by=By.CSS_SELECTOR, value='a').get_attribute('href')
#         if 'cafe' in chkCafe:
#             untilEleGone(item, '.shs_list')
#             break

#     myCafeGo = searchElement('.mycafe .btn_cafe_more',driver)
#     untilEleGone(myCafeGo[0], '.mycafe')

#     myCafeList = searchElement('.list_cafe__favorites li',driver)
#     with open("./etc/cafe_info.txt", "r") as f:
#         getCafeNameList = f.readlines()
#         getCafeName = getCafeNameList[0]
#         getCafeName = getCafeName.replace(" ", "")

#     for onCafe in myCafeList:
#         chkCafeTitle = onCafe.find_element(
#             by=By.CSS_SELECTOR, value='.title').text
#         chkCafeTitle = chkCafeTitle.replace(" ", "")

#         if chkCafeTitle in getCafeName:
#             untilEleGone(onCafe, '.list_cafe__favorites')
#             break

#     # 카페 진입 끝


def getBlogContentChrome(subjectArr,driver):
   
    with open('./etc/find_keyword.txt', 'r') as r:
        allKeyword = r.readlines()
        
    driver.get('https://www.google.com/')
    
    
    searchCount = 0
    while True:
        searchCount += 1
        keyCount = random.randrange(0, len(allKeyword))
        getKeyword = allKeyword[keyCount]
        getKeyword = getKeyword.replace('\n', '')
        searchBar = searchElement_seterr('.gLFyf',driver)
        
        if searchBar == 'error':
            return 'error'
        
        wait_float(0.5,0.8)
        searchBar[-1].click()
        wait_float(0.5,0.8)
        searchBar[-1].clear()
        wait_float(0.5,0.8)
        searchBar[-1].send_keys(f'site:blog.naver.com {getKeyword}')
        pg.press('enter')
        wait_float(0.8,1.5)
        pg.press('enter')
        
        nowpage = 0
        if searchCount == 1:
            getTools = searchElement('.t2vtad',driver)
            getTools[0].click()
            wait_float(0.3,0.9)
            
            getToolsIf = searchElement('.KTBKoe',driver)
            for getToolsIfOn in getToolsIf:
                if '날짜' in getToolsIfOn.text:
                    getToolsIfOn.click()
            wait_float(0.3,0.9)
                    
            getPeriodIf = searchElement('.y0fQ9c',driver)
            for getPeriodIfOn in getPeriodIf:
                if '설정' in getPeriodIfOn.text:
                    getPeriodIfOn.click()
            wait_float(0.3,0.9)
            
            today = datetime.today()
            print(today)
            
            before_one_year = today - relativedelta(years=3)

            # this_month_first = datetime(before_one_year.year, before_one_year.month, 1)
            # chkMonthFirst = this_month_first.strftime('%m/%d/%Y')
            # driver.find_element(by=By.CSS_SELECTOR, value='.OouJcb').send_keys(chkMonthFirst)
            # next_month = datetime(before_one_year.year, before_one_year.month, 1) + relativedelta(months=1)
            # this_month_last = next_month + relativedelta(seconds=-1)
            # chkMonthLast = this_month_last.strftime('%m/%d/%Y')
            
            chkThreeyearAgo = before_one_year.strftime('%m/%d/%Y')
            driver.find_element(by=By.CSS_SELECTOR, value='.OouJcb').send_keys(chkThreeyearAgo)
            driver.find_element(by=By.CSS_SELECTOR, value='.rzG2be').send_keys(chkThreeyearAgo)
            wait_float(0.8,1.5)
            pg.press('enter')
            
        try:
            driver.find_element(by=By.CSS_SELECTOR, value='.NVbCr')
        except:
            continue
        getPagingList = searchElement('.NVbCr',driver)
        if len(getPagingList) < 2:
            continue
        getPgCount = random.randrange(0,len(getPagingList))
        if getPgCount != nowpage:
            nowpage = getPgCount
            getPagingList[getPgCount].click()
        

        
        linkSearchCount = 0
        resetContent = ''
        while True:
            wait_float(0.5,1.2)
            linkSearchCount += 1
            if linkSearchCount > 10:
                resetContent = "on"
                break
            try:
                getBlogLink = driver.find_elements(by=By.CSS_SELECTOR, value='.yuRUbf')
                if len(getBlogLink) > 3:
                    break
                
            except:
                pass
        if resetContent == 'on':
            continue
            
                
            
            
        
        getBlogLinkCount = random.randrange(0,len(getBlogLink))
        
        getInfoPostLink = getBlogLink[getBlogLinkCount].find_element(by=By.CSS_SELECTOR, value='a').get_attribute('href')
        if '//m.' not in getInfoPostLink:
            getInfoPostLink = getInfoPostLink.replace('//', '//m.')


        page = requests.get(getInfoPostLink)
        soup = bs(page.text, "html.parser")
        elements = soup.select('.se-module.se-module-text')
        
        # se-title-text 제목
        
        getSubjectEle = soup.select('.se-title-text')
        try:
            getSubject = str(getSubjectEle[0])
        except:
            continue
        
        getSubject = getSubject.split('-->')[-2].replace('','').split('<!--')[0]
        getSubject = re.sub(r"[^\uAC00-\uD7A3\s]", "", getSubject)
        getSubjectArr = getSubject.replace('  ', ' ').split(' ')

        allStr = []
        for ele in elements:
            p = re.compile('[\uAC00-\uD7A30-9a-zA-Z\s]+')
            chkResult = p.findall(str(ele))
            allStr = allStr + chkResult

        p_str = re.compile(r'[a-zA-Z0-9,|\n]+')
        p_space = re.compile('\s\s')


        for i in range(1, len(allStr)):
            for j, strin in enumerate(allStr):
                getStr = p_str.search(strin)
                if getStr is not None:
                    allStr.pop(j)
                    break
                getSpace = p_space.search(strin)
                if getSpace is not None:
                    allStr.pop(j)
                    break
                if strin == " ":
                    allStr.pop(j)
                    break
        
        getAllStr = []
        for strOn in allStr:
            addVal = ''
            for tempsub in getSubjectArr:
                if len(tempsub) < 2:
                    continue
                if tempsub in strOn:
                    addVal = ''
                    break
                addVal = 'on'
            
            if addVal == 'on':
                getAllStr.append(strOn)
        
        
        
        
        allStr = " ".join(getAllStr)
        
        if len(allStr) < 400:
            continue
        if len(allStr) > 600:
            sliceRanNum = random.randrange(350, 450)
            allStr = allStr[0:sliceRanNum]
            
        # 제목에 들어간 단어들 삭제하기
        break

    resetStrArr = allStr.split(' ')

    resetListArr = list_chunk(resetStrArr, 12)
    for resetList in resetListArr:
        setRan = random.randrange(2, 5)
        resetOn = random.sample(range(1, 13), setRan)

        if resetList == "":
            continue

        for inon in resetOn:
            changeRanCount = random.randrange(0, len(subjectArr))
            chkChangeRan = random.randrange(1, 4)
            if chkChangeRan == 1:
                try:
                    resetList[inon - 1] = subjectArr[changeRanCount]
                except:
                    pass
            else:
                try:
                    resetList[inon - 1] = ''
                except:
                    pass

    imgLineCountBasic = divmod(len(resetListArr), 2)
    imgLineCount = random.randrange(
        int(imgLineCountBasic[0]) - 4, int(imgLineCountBasic[0]) + 4)

    allContent = ''
    for i, setList in enumerate(resetListArr):
        if imgLineCount == i:
            allContent = allContent + 'img_line|randomimg\n'
            insubject = " ".join(subjectArr)
            allContent = allContent + insubject + '\n'
        for setStr in setList:
            if setStr == '':
                continue
            elif len(setStr) > 20:
                continue
            allContent = allContent + setStr
            allContent = allContent + ' '
        
        if len(resetListArr)-1 != i:
            allContent = allContent + '\n'
        
    allContent = allContent + subjectArr[-1]

    driver.close()
    return allContent
        
        
        

    
    
    # #시작일
    # OouJcb
    
    # #종료일
    # rzG2be
    
    
    
    
    
    


def getSubjectArrToCafe(driver):
    with open('./etc/subject_cafe_list.txt', 'r') as r:
        cafeList = r.readlines() 
    
    onCafe = cafeList[random.randrange(0,len(cafeList))]
    onCafe = onCafe.replace('\n','')
    driver.get(onCafe)
    searchElement('.list_area', driver)
    
    while True:
        boardBox = driver.find_elements(by=By.CSS_SELECTOR, value=".board_box")
        before_one_day = datetime.now() - timedelta(days=2)
        
        setTime = boardBox[-1].find_element(by=By.CSS_SELECTOR, value=".time")
        if ':' in setTime.text:
            pg.press('end')
            wait_float(1.5,2.5)
            continue
        setTime = '20' + setTime.text[:-1].replace('.','-')
        getBoardTime = datetime.strptime(setTime, '%Y-%m-%d')
        
        if getBoardTime < before_one_day:
            break
        else:
            pg.press('end')
            wait_float(1.5,2.5)

    
    
    titleList = driver.find_elements(by=By.CSS_SELECTOR, value=".board_box .tit")
    while True:
        getRanTitNum = random.randrange(0,len(titleList))

        
        if len(titleList[getRanTitNum].text) < 13:
            del titleList[getRanTitNum]
            continue
        
        chkSubjectArr = ['핸드','휴대','휴싸방','좌표어때','가입','등업','안녕','하이','댓글']
        for chkOn in chkSubjectArr:
            if chkOn in titleList[getRanTitNum].text:
                del titleList[getRanTitNum]
                continue
        
        getSubjectOn = titleList[getRanTitNum].text
        break
    getSubject = re.sub(r"[^\uAC00-\uD7A30-9a-zA-Z\s]", "", getSubjectOn)
    subjectArr = getSubject.split(' ')
    
    return subjectArr
    
    
    
    
     
    

def getBlogContent(subjectArr):
    # 블로그 글따기 시작

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    driver.get('https://section.blog.naver.com/BlogHome.naver')

    try:
        popup = driver.find_element(
            by=By.CSS_SELECTOR, value='#floatingda_home')
        popupClostBtn = popup.find_elements(by=By.CSS_SELECTOR, value='button')
        popupClostBtn[-1].click()
    except:
        pass

    while True:

        try:
            nCategoryList = driver.find_elements(
                by=By.CSS_SELECTOR, value='.navigator_category a')
            categoryRanVal = random.randrange(0, len(nCategoryList) - 1)
            nCategoryList[categoryRanVal].click()

            wait_float(0.5, 0.9)

            paginationNum = driver.find_elements(
                by=By.CSS_SELECTOR, value='.pagination span')
            driver.execute_script(
                "arguments[0].scrollIntoView();", paginationNum[0])
            paginationRanVal = random.randrange(0, len(paginationNum) - 1)
            getClickPage = paginationNum[paginationRanVal].find_element(
                by=By.CSS_SELECTOR, value='a')
            getClickPage.click()
            wait_float(0.5, 0.9)

            infoPostList = driver.find_elements(
                by=By.CSS_SELECTOR, value='.info_post')
            infoPostRanVal = random.randrange(0, len(infoPostList) - 1)
            getInfoPostTag_a = infoPostList[infoPostRanVal].find_element(
                by=By.CSS_SELECTOR, value='.desc a')
            getInfoPostLink = getInfoPostTag_a.get_attribute('href')
            getInfoPostLink = getInfoPostLink.replace('//', '//m.')
        except:
            driver.refresh()
            focus_window('blog.naver')
            pg.press('F5')
            wait_float(2.5, 3.5)
            continue

        page = requests.get(getInfoPostLink)
        soup = bs(page.text, "html.parser")
        elements = soup.select('.se-module.se-module-text')

        allStr = []
        for ele in elements:
            p = re.compile('[\uAC00-\uD7A30-9a-zA-Z\s]+')
            chkResult = p.findall(str(ele))
            allStr = allStr + chkResult

        p_str = re.compile(r'[a-zA-Z0-9,|\n]+')
        p_space = re.compile('\s\s')

        for i in range(1, len(allStr)):
            for j, strin in enumerate(allStr):
                getStr = p_str.search(strin)
                if getStr is not None:
                    allStr.pop(j)
                    break
                getSpace = p_space.search(strin)
                if getSpace is not None:
                    allStr.pop(j)
                    break
                if strin == " ":
                    allStr.pop(j)
                    break
        allStr = "".join(allStr)
        if len(allStr) < 600:
            continue
        if len(allStr) > 1200:
            sliceRanNum = random.randrange(1050, 1150)
            allStr = allStr[0:sliceRanNum]
        break

    resetStrArr = allStr.split(' ')

    resetListArr = list_chunk(resetStrArr, 12)
    for resetList in resetListArr:
        setRan = random.randrange(2, 5)
        resetOn = random.sample(range(1, 13), setRan)

        if resetList == "":
            continue

        for inon in resetOn:
            changeRanCount = random.randrange(0, len(subjectArr))
            chkChangeRan = random.randrange(1, 6)
            if chkChangeRan == 1:
                try:
                    resetList[inon - 1] = subjectArr[changeRanCount]
                except:
                    pass
            else:
                try:
                    resetList[inon - 1] = ''
                except:
                    pass

    imgLineCountBasic = divmod(len(resetListArr), 2)
    imgLineCount = random.randrange(
        int(imgLineCountBasic[0]) - 4, int(imgLineCountBasic[0]) + 4)

    allContent = ''
    for i, setList in enumerate(resetListArr):
        if imgLineCount == i:
            allContent = allContent + 'img_line|randomimg\n'
        for setStr in setList:
            if setStr == '':
                continue
            elif len(setStr) > 20:
                continue
            allContent = allContent + setStr
            allContent = allContent + ' '
        allContent = allContent + '\n'

    driver.close()
    return allContent


# subjectArr
def list_chunk(lst, n):
    return [lst[i:i+n] for i in range(0, len(lst), n)]
