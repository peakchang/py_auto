import random
import threading
import time
from datetime import datetime, timedelta
import sys
import os
from pathlib import Path
from typing import Optional
# from pyparsing import And
import requests
from bs4 import BeautifulSoup as bs
import json
import re
import pyautogui as pg
import pyperclip
import pygetwindow as gw
import clipboard as cb
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import WebDriverException

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import requests
import winsound as ws
import glob
import asyncio








def writeBlog(driver,getDict,goChk):
    navItem = searchElement('.nav_item', driver)
    for mitem in navItem:
        if mitem.text == '블로그':
            mitem.click()
            break
    
    dirList = os.listdir(f"{os.getcwd()}\\content")
    
    lastForIdx = len(dirList)
    addHour = 0
    for idx,dir in enumerate(dirList):
        
        if goChk == 0:
            addHour += 2
        else:
            if idx != 0:
                addHour += 2
        menu_my_blog = searchElement('.menu_my_blog .item', driver)
        menu_my_blog[1].click()
        
        driver.switch_to.window(driver.window_handles[1])
        
        if getDict['middleVal'] == 0:
            pg.alert(f'글쓰기를 시작합니다!!')
        # driver.to_switch()
        driver.switch_to.window(driver.window_handles[1])
        
        
        
        
        
        driver.switch_to.frame('mainFrame')
        
        focus_window('블로그')
        writeArea = searchElement('.se-component-content',driver)
        wait_float(2.5,3.5)
        
        try:
            preWritePopup = driver.find_element(by=By.CSS_SELECTOR, value=".se-popup-button-cancel")
            preWritePopup.click()
            wait_float(0.5,0.9)
            
        except:
            pass
        
        subject_area = writeArea[0].find_element(by=By.CSS_SELECTOR, value=".se-section-documentTitle")
        subject_area.click()
        
        try:
            with open(f'./content/{dir}/content.txt', 'rt', encoding='UTF8') as f:
                getLines = f.readlines()
        except:
            with open(f'./content/{dir}/content.txt', 'r') as f:
                getLines = f.readlines()
            
        for i, getline in enumerate(getLines):
            focus_window('블로그')
            getline = getline.replace('\n', '')
            chkImg = getline.split('|')
            if chkImg[0] == 'img_line':
                nowPath = os.getcwd()
                
                img_btn = searchElement('.se-image-toolbar-button', driver)
                img_btn[0].click()
                wait_float(1.5,2.3)
                
                imagePath = nowPath + f"\content\{dir}"
                wait_float(1.5, 2.2)
                pyperclip.copy(imagePath)
                wait_float(0.5, 0.9)
                pg.hotkey('ctrl','v')
                wait_float(0.5, 0.9)
                pg.press('enter')
                
                wait_float(0.9, 1.6)
                pyperclip.copy(chkImg[1])
                wait_float(0.5, 0.9)
                pg.hotkey('ctrl','v')
                wait_float(0.5, 0.9)
                pg.press('enter')
                wait_float(3.5,4.5)
                continue
            
            if i == 0:
                writeArea[0].click()
                keyboard.write(text=getline, delay=0.05)
                wait_float(1.2,2.8)
            elif i == 1:
                writeArea[1].click()
                keyboard.write(text=getline, delay=0.05)
                wait_float(0.5,0.9)
                pg.press('enter')
                wait_float(0.5,0.9)
            elif getline == 'enter':
                wait_float(0.5,0.9)
                pg.press('enter')
                wait_float(0.5,0.9)
            else:
                keyboard.write(text=getline, delay=0.05)
                wait_float(0.5,0.9)
                pg.press('enter')
                wait_float(0.5,0.9)
                
        try:
            helpCloseBtn = driver.find_element(by=By.CSS_SELECTOR, value=".se-help-panel-close-button")
            helpCloseBtn.click()
        except:
            pass
        
        publichBtn = searchElement('.publish_btn__Y5mLP',driver)
        publichBtn[0].click()
        wait_float(1.5,2.5)
        tagArea = searchElement('.tag_textarea__iAnXk',driver)
        tagArea[0].click()
        
        try:
            with open(f'./content/{dir}/tag_list.txt', 'rt', encoding='UTF8') as tagr:
                tagList = tagr.readlines()
        except:
            with open(f'./content/{dir}/tag_list.txt', 'r') as tagr:
                tagList = tagr.readlines()
                    
                    
        if tagList is not []:
            for tag in tagList:
                writeTag = tag.replace('\n', '')
                keyboard.write(text=writeTag, delay=0.05)
                wait_float(0.5,1.2)
                pg.press('enter')
        
        if goChk == 0:
            chkVal = pg.confirm(text='글을 지금 바로 등록하시겠습니까?', buttons=['now','reserve'])
            if chkVal == 'now':
                confirmBtn = searchElement('.confirm_btn__Dv9du',driver)
                confirmBtn[0].click()
            else:
                wait_float(0.5,0.9)
                reserveRadio = searchElement('.radio_label__wZWth',driver)
                reserveRadio[5].click()
                wait_float(0.5,0.9)
                now = datetime.now()
                setHour = now.hour + addHour
                setHourStr = str(setHour)
                planService = Select(driver.find_element(by=By.CSS_SELECTOR, value=f'.hour_option__XigHn'))
                planService.select_by_visible_text(setHourStr)
                pg.alert('예약 조건 확인 후 글 발행하고 엔터!!!')
        else:
            if idx == 0:
                confirmBtn = searchElement('.confirm_btn__Dv9du',driver)
                confirmBtn[0].click()
                wait_float(3.5,5.5)
            else:
                wait_float(0.5,0.9)
                reserveRadio = searchElement('.radio_label__wZWth',driver)
                reserveRadio[5].click()
                wait_float(0.5,0.9)
                now = datetime.now()
                setHour = now.hour + addHour
                setHourStr = str(setHour)
                planService = Select(driver.find_element(by=By.CSS_SELECTOR, value=f'.hour_option__XigHn'))
                planService.select_by_visible_text(setHourStr)
                pg.alert('예약 조건 확인 후 글 발행하고 엔터!!!')
        
        wait_float(3.5,5.5)
        
        if idx == lastForIdx - 1:
            try:
                miniPopup = driver.find_element(by=By.CSS_SELECTOR, value="#floatingda_content")
                miniPopupClose = miniPopup.find_element(by=By.CSS_SELECTOR, value="button")
                driver.execute_script("arguments[0].scrollIntoView();", miniPopupClose)
                wait_float(0.5,1.2)
                miniPopupClose.click()
                wait_float(0.5,1.2)
            except:
                pass
            
            getUrl = searchElement('._transPosition',driver)
            getUrl[0].click()
            wait_float(1.5,2.5)
            
        pg.press('enter')
        driver.switch_to.default_content()
        while True:
            if len(driver.window_handles) > 1:
                driver.switch_to.window(driver.window_handles[1])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
            else:
                break
            wait_float(0.5,0.9)
    
    goToNaverMain = searchElement('.link_naver',driver)
    goToNaverMain[0].click()
    driver.get('https://www.naver.com')
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            


def allowListVisit(driver):
    
    navItem = searchElement('.nav_item',driver)
    for mitem in navItem:
        if mitem.text == '블로그':
            mitem.click()
            break
    
    menu_my_blog = searchElement('.menu_my_blog .item',driver)
    menu_my_blog[0].click()
    
    driver.switch_to.window(driver.window_handles[1])
    driver.switch_to.frame('mainFrame')
    
    wait_float(1.5,2.5)
    
    try:
        closePopupBtn = driver.find_element(by=By.CSS_SELECTOR, value="#not_see")
        closePopupBtn.click()
    except:
        pass
    
    try:
        closePopupBtn = driver.find_element(by=By.CSS_SELECTOR, value=".popup_da_btn_area ._btn_close")
        closePopupBtn.click()
    except:
        pass
    
    getUrl = searchElement('._transPosition',driver)
    getUrl[0].click()
    
    wait_float(1.5,2.5)
    pg.press('enter')
    
    nowBlogLink = pyperclip.paste()
    nowBlogLinkSplit = nowBlogLink.split('/')
    
    openVisitListBtn = searchElement(f'#Sympathy{nowBlogLinkSplit[-1]} .bu_arr',driver)
    openVisitListBtn[0].click()
    
    
    wait_float(1.5,2.5)
    
    driver.switch_to.frame(f'sympathyFrm{nowBlogLinkSplit[-1]}')
    
    visitListWrap = searchElement('.wrap_blog2_sympathy',driver)
    try:
        visitList = visitListWrap[0].find_elements(by=By.CSS_SELECTOR, value=".nick")
    except:
        driver.close()
        wait_float(0.3,0.9)
        driver.switch_to.window(driver.window_handles[0])
        wait_float(0.3,0.9)
        driver.switch_to.default_content()
        
        goToMain = searchElement('.link_naver',driver)
        goToMain[0].click()
        
        while True:
            try:
                WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#query")))
                break
            except:
                driver.get('https://www.naver.com')
                
        return
     
    
    
    for visitCount in range(len(visitList)):
        
        driver.switch_to.default_content()
        wait_float(0.3,0.9)
        driver.switch_to.frame('mainFrame')
        wait_float(0.3,0.9)
        driver.switch_to.frame(f'sympathyFrm{nowBlogLinkSplit[-1]}')
        wait_float(0.3,0.9)
        visitList = searchElement('.wrap_blog2_sympathy .nick',driver)
        print(visitList[visitCount].text)
        visitList[visitCount].click()
        wait_float(0.7,1.2)
        driver.switch_to.window(driver.window_handles[2])
        
        try:
            wait_float(0.7,1.2)
            driver.switch_to.default_content()
            driver.switch_to.frame('mainFrame')
            
            blogMenuChk = searchElement('#blog-menu .menu1 li a',driver)
            if(len(blogMenuChk) > 1):
                blogMenuChk[1].click()
                searchElement('#blog-menu',driver)
            
            
            
            try:
                postListOpenBtn = driver.find_element(by=By.CSS_SELECTOR, value="#toplistSpanBlind")
            except:
                wait_float(0.5,1.5)
                driver.close()
                wait_float(0.3,0.9)
                driver.switch_to.window(driver.window_handles[1])
                wait_float(0.3,0.9)
                continue
            
            
            # 여기서 블로그 말고 프롤로그면 블로그 클릭하게 하기
            
            while True:
                print(postListOpenBtn.text)
                if postListOpenBtn.text == '목록닫기':
                    break
                else:
                    wait_float(0.5,1.3)
                    try:
                        postListOpenBtn.click()
                    except:
                        wait_float(2.5,3.5)
                        continue
            wait_float(1.5,2.5)
            postList = searchElement('.blog2_categorylist',driver)
            
            for getPostLinkCount in range(3):
                try:
                    getPostLink = postList[getPostLinkCount].find_element(by=By.CSS_SELECTOR, value=".ell2")
                    break
                except:
                    pass
            
            getPostLink.click()
            wait_float(1.3,2.5)
            
            try:
                gongamBtn = driver.find_element(by=By.CSS_SELECTOR, value='.u_likeit_list_btn')
                getGonggamStatus = gongamBtn.get_attribute('aria-pressed')
                print(getGonggamStatus)
                if getGonggamStatus == 'false':
                    gongamBtn = searchElement('.u_ico',driver)
                    wait_float(0.3,0.9)
                    gongamBtn[-1].click()
                    wait_float(3.2,4.5)
                else:
                    pass
            except:
                pass

            
            # gongamBtn[1].click()
            wait_float(0.5,1.5)
            driver.close()
            wait_float(0.3,0.9)
            driver.switch_to.window(driver.window_handles[1])
            wait_float(0.3,0.9)
        except:
            wait_float(0.5,1.5)
            driver.close()
            wait_float(0.3,0.9)
            driver.switch_to.window(driver.window_handles[1])
            wait_float(0.3,0.9)
        
        
    driver.close()
    wait_float(0.3,0.9)
    driver.switch_to.window(driver.window_handles[0])
    wait_float(0.3,0.9)
    driver.switch_to.default_content()
    
    goToMain = searchElement('.link_naver',driver)
    goToMain[0].click()
    
    while True:
        try:
            WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#query")))
            break
        except:
            driver.get('https://www.naver.com')
            
    
def getBlogContent(getVal):
    
    global driver
    getLink = getVal['getText']
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    driver.get(getLink)
    
    contentBox = searchElement('.tt_article_useless_p_margin p',driver)
    
    allContent = ''
    for contentLine in contentBox:
        if contentLine is None or contentLine == '':
            continue
        
        allContent = allContent + contentLine.text + '\n'
    
    with open('./getblogcontent.txt', 'w') as f:
        f.write(allContent)
        
    driver.quit()
    sys.exit(0)
    
    
    
    
    


def blogRankChk(getDict):
    getInfoPostLink = getDict['getText']
    exLineNum = getDict['nlist']
    wb = load_workbook('./etc/nid.xlsx')
    ex = wb.active
    searchId = ex.cell(exLineNum, 1).value
    
    global driver

    
    if getInfoPostLink == "":
        pg.alert('검색어를 입력하세요!')
        return
        
    
    options = Options()
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, chrome_options=options)
    driver.get('https://www.naver.com')
    
    serchBar = searchElement('#query',driver)
    serchBar[0].send_keys(getInfoPostLink)
    wait_float(0.3,0.8)
    pg.press('enter')
    wait_float(0.3,0.8)
    
    lnbMenu = searchElement('.lnb_menu .menu',driver)
    
    for menu in lnbMenu:
        if 'VIEW' in menu.text:
            menu.click()
            break
        
    blogChk = searchElement('.type_sort a',driver)
    blogChk[1].click()
    
    listCount = 0
    while True:
        
        if listCount % 20 == 0:
            pg.press('end')
        
        allList = searchElement('.lst_total li',driver)
        listCount += 1
        
        try:
            getHref = allList[listCount].find_element(by=By.CSS_SELECTOR, value='.thumb_single')
            print(getHref.get_attribute('href'))
            
        except:
            continue
        
        if searchId in getHref.get_attribute('href'):
            pg.alert(f'현재 {listCount}번째 있습니다~')
            exitApp()
        
    
     
    
    
    
    
    





def blogReplyReady(getValList):
    
    if getValList['nlist'] == 1:
        pg.alert('아이디가 선택되지 않았습니다. 다시 실행해주세요')
        exitApp()
    
    global driver
    
    exLineNum = getValList['nlist']
    wb = load_workbook('./etc/nid.xlsx')
    ex = wb.active
    
    options = Options()
    user_data = 'C:\\Users\\pcy\\AppData\\Local\\Google\\Chrome\\User Data\\default'
    service = Service(ChromeDriverManager().install())
    options.add_argument(f"user-data-dir={user_data}")
    options.add_argument(f'--profile-directory={ex.cell(exLineNum, 3).value}')
    driver = webdriver.Chrome(service=service, chrome_options=options)
    
    driver.get('https://www.naver.com')
    loginBtn = searchElement('.sc_login',driver)
    loginBtn[0].click()
    
    # while True:
    
    searchElement('#id',driver)
    focus_window('chrome')
    wait_float(0.3,0.9)
    while True:
        
        pg.click(200,500)
        inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
        inputId.click()
        wait_float(0.3,0.9)
        cb.copy(ex.cell(exLineNum, 1).value)
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'a')
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'v')
        inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
        if inputId.get_attribute('value') != "":
            break
        
    while True:
        inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
        inputPw.click()
        wait_float(0.3,0.9)
        cb.copy(ex.cell(exLineNum, 2).value)
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'a')
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'v')
        inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
        if inputPw.get_attribute('value') != "":
            break
    
    btnLogin = searchElement('.btn_login',driver)
    btnLogin[0].click()
    
    # 블로그 링크따기
    
    while True:
        try:
            WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#query")))
            break
        except:
            driver.get('https://www.naver.com')
    
    navItem = searchElement('.nav_item',driver)
    for mitem in navItem:
        if mitem.text == '블로그':
            mitem.click()
            break
    
    menu_my_blog = searchElement('.menu_my_blog .item',driver)
    menu_my_blog[0].click()
    
    driver.switch_to.window(driver.window_handles[1])
    driver.switch_to.frame('mainFrame')
    
    wait_float(1.5,2.5)
    
    try:
        closePopupBtn = driver.find_element(by=By.CSS_SELECTOR, value="#not_see")
        closePopupBtn.click()
    except:
        pass
    
    try:
        closePopupBtn = driver.find_element(by=By.CSS_SELECTOR, value=".popup_da_btn_area ._btn_close")
        closePopupBtn.click()
    except:
        pass
    
    getUrl = searchElement('._transPosition',driver)
    getUrl[0].click()
    wait_float(1.5,2.5)
    pg.press('enter')
    driver.switch_to.default_content()
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    
    goToNaverMain = searchElement('.link_naver',driver)
    goToNaverMain[0].click()
    
    # 블로그 링크 따기 끝~~~
    
    blogReplyWork()




def blogReplyWork(driver):
    navItem = searchElement('.nav_item',driver)
    for mitem in navItem:
        if mitem.text == '카페':
            mitem.click()
            break
    
    cafeList = searchElement('.user_mycafe_info',driver)
    getInCafe = ""
    for cafeOn in cafeList:
        if "소셜공간" in cafeOn.text:
            cafeOn.click()
            getInCafe = "on"
            break
    
    if getInCafe == "":
        driver.get('https://cafe.naver.com/sens3')
    else:
        driver.switch_to.window(driver.window_handles[0])
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        
    cafeWriteBtn = searchElement('.cafe-write-btn',driver)
    
    if "가입" in cafeWriteBtn[0].text:
        pg.alert('카페 가입하기~~')
    
    workBoardLink = searchElement('#menuLink226',driver)
    workBoardLink[0].click()
    
    driver.switch_to.frame('cafe_main')
    
    
    
    # 카페에 글 작성하기
    
    workBoardWriteBtn = searchElement('#writeFormBtn',driver)
    workBoardWriteBtn[0].click()
    
    wait_float(1.5,2.5)
    driver.switch_to.window(driver.window_handles[1])
    
    
    subjectArea = searchElement('.FlexableTextArea',driver)
    subjectArea[0].click()
    
    with open('./etc/social_cafe_content.txt', 'r') as r:
        cafeContent = r.readlines()
    
    keyboard.write(text=cafeContent[0], delay=0.05)
    wait_float(0.3,0.9)
    
    contentArea = searchElement('.se-content',driver)
    contentArea[0].click()
    wait_float(0.3,0.9)
    pg.hotkey('ctrl', 'a')
    
    for i,conLine in enumerate(cafeContent):
        if i == 0:
            continue
        keyboard.write(text=conLine, delay=0.03)
        wait_float(0.5,1.5)
    pg.press('enter')
    pg.hotkey('ctrl', 'v')
    wait_float(1.5,2.5)
    
    BaseButton = searchElement('.BaseButton',driver)
    BaseButton[0].click()
    
    
    
    
    wait_float(1.5,2.5)
    driver.switch_to.frame('cafe_main')
    
    wait_float(0.5,0.9)
    driver.close()
    
    driver.switch_to.window(driver.window_handles[0])
    
    
    
    workCafeLink = pyperclip.paste()
    workCafeNum = workCafeLink.split('/')[-1]
    preNick = ""
    
    # 카페에 글 작성하기 끝~
    
    
    
    print(workCafeLink)
    
    forVal = random.randrange(6,8)
    
    ici = 0
    while True:
        
        
        ici += 1
        driver.switch_to.default_content()
        
        wait_float(0.3,0.9)
        workBoardLink = searchElement('#menuLink226',driver)
        workBoardLink[0].click()
        
        wait_float(0.3,0.9)
        
        driver.switch_to.frame('cafe_main')
        
        wait_float(0.3,0.9)
        
        articleDiff = searchElement('.article-board',driver)
        articleList = articleDiff[1].find_elements(by=By.CSS_SELECTOR, value=".td_article")
        
        
        
        if str(workCafeNum) in articleList[ici].find_element(by=By.CSS_SELECTOR, value=".board-number").text:
            forVal = forVal + 1
            continue
        
        wait_float(0.3,0.9)
        clickArticleTarget = articleList[ici].find_element(by=By.CSS_SELECTOR, value=".article")
        wait_float(0.3,0.9)
        # clickArticleTarget.click()
        untilEleShow(clickArticleTarget, '.nickname')
        
        
        nickname = searchElement('.nickname',driver)
        if preNick == nickname[0].text:
            driver.back()
            wait_float(2.1,3.7)
            continue
        else:
            preNick = nickname[0].text
        
        chkLinkTag = driver.find_elements("xpath", "//*[contains(@class, 'se-fs-')]")

        for chkLink in chkLinkTag:
            try:
                getOtherBlogLink = chkLink.find_element(by=By.CSS_SELECTOR, value="a").get_attribute('href')
                if 'blog' in str(getOtherBlogLink):
                    chkOtherBlogLink = getOtherBlogLink.split('/')
                    if len(chkOtherBlogLink) < 5:
                        forVal = forVal + 1
                        driver.back()
                        wait_float(2.1,3.7)
                        break
                    else:
                        
                        chkLink.click()
                        
                        if 'm.' in getOtherBlogLink:
                            wait_float(0.3,0.9)
                            driver.switch_to.window(driver.window_handles[1])
                            driver.switch_to.default_content()
                            gongamBtn = searchElement('.u_ico',driver)
                            wait_float(0.3,0.9)
                            gongamBtn[-1].click()
                            wait_float(2.2,2.9)
                            driver.close()
                            wait_float(0.3,0.9)
                            driver.switch_to.window(driver.window_handles[0])
                            wait_float(0.3,0.9)
                            
                        else:
                            wait_float(0.3,0.9)
                            driver.switch_to.window(driver.window_handles[1])
                            driver.switch_to.default_content()
                            wait_float(0.3,0.9)
                            driver.switch_to.frame('mainFrame')
                            wait_float(0.3,0.9)
                            gongamBtn = searchElement('.u_ico',driver)
                            gongamBtn[1].click()
                            wait_float(2.2,2.9)
                            driver.switch_to.default_content()
                            wait_float(0.3,0.9)
                            driver.close()
                            wait_float(0.3,0.9)
                            driver.switch_to.window(driver.window_handles[0])
                            wait_float(0.3,0.9)
                        
                        driver.switch_to.frame('cafe_main')
                        replyArea = searchElement('.comment_inbox_text',driver)
                        replyArea[0].click()
                        
                        for i,conLine in enumerate(cafeContent):
                            if i == 0:
                                continue
                            keyboard.write(text=conLine, delay=0.03)
                            wait_float(0.5,1.5)
                        pg.press('enter')
                        pg.hotkey('ctrl', 'v')
                        wait_float(1.5,2.5)
                        
                        replySuccessBtn = searchElement('.btn_register',driver)
                        driver.execute_script("arguments[0].scrollIntoView();", replySuccessBtn[0])
                        replySuccessBtn[0].click()
                        wait_float(5.5,7.5)
                        
                        driver.back()
                        break
            except:
                pass
        if ici >= forVal:
            break









# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>함수 시작염

# 상품 들어가서 스크롤 내리고 나오기

def makeBlogContent(getInfoPostLink):
    
    if '//m.' not in getInfoPostLink:
        getInfoPostLink = getInfoPostLink.replace('//', '//m.')
    
    print(type(getInfoPostLink))

    page = requests.get(getInfoPostLink)
    soup = bs(page.text, "html.parser")
    elements = soup.select('.se-module.se-module-text')
    
    
    allStr = []
    chkCount = 0
    for ele in elements:
        chkCount += 1
        p = re.compile('[\uAC00-\uD7A30-9a-zA-Z\s]+')
        chkResult = p.findall(str(ele))
        if chkCount == 1:
            subjectTemp = chkResult
        allStr = allStr + chkResult

    p_str = re.compile(r'[a-zA-Z0-9,|\n]+')
    p_space = re.compile('\s\s')
    
    subjectArrTemp = []
    for sentence_s in subjectTemp:
        getStr = p_str.search(sentence_s)
        if getStr is None:
            subjectArrTemp.append(sentence_s)
    
    
    
    tempSubjectOn = ''
    for tempss in subjectArrTemp:
        if tempss == ' ':
            continue
        tempSubjectOn = tempSubjectOn + tempss
    
    
    
    subjectArr = tempSubjectOn.split(' ')
        
        

    for i in range(1, len(allStr)):
        for j, strin in enumerate(allStr):
            getStr = p_str.search(strin)
            if getStr is not None:
                allStr.pop(j)
                break
            getSpace = p_space.search(strin)
            if getSpace is not None:
                allStr.pop(j)
                break
            if strin == " ":
                allStr.pop(j)
                break
    allStr = "".join(allStr)
    
    # if len(allStr) < 1500:
    #     continue
    # if len(allStr) > 1900:
    #     sliceRanNum = random.randrange(1050, 1150)
    #     allStr = allStr[0:sliceRanNum]
    # break

    resetStrArr = allStr.split(' ')

    resetListArr = list_chunk(resetStrArr, 12)
    for resetList in resetListArr:
        setRan = random.randrange(2, 5)
        resetOn = random.sample(range(1, 13), setRan)

        if resetList == "":
            continue

        for inon in resetOn:
            changeRanCount = random.randrange(0, len(subjectArr))
            chkChangeRan = random.randrange(1, 6)
            if chkChangeRan == 1:
                try:
                    resetList[inon - 1] = subjectArr[changeRanCount]
                except:
                    pass
            else:
                try:
                    resetList[inon - 1] = ''
                except:
                    pass

    imgLineCountBasic = divmod(len(resetListArr), 2)
    imgLineCount = random.randrange(
        int(imgLineCountBasic[0]) - 4, int(imgLineCountBasic[0]) + 4)

    allContent = ''
    for i, setList in enumerate(resetListArr):
        if imgLineCount == i:
            allContent = allContent + 'img_line|randomimg\n'
        for setStr in setList:
            if setStr == '':
                continue
            elif len(setStr) > 20:
                continue
            allContent = allContent + setStr
            allContent = allContent + ' '
        allContent = allContent + '\n'

    # driver.close()
    pg.alert(allContent)
    with open('./etc/text.txt', 'w') as f:
        f.write(allContent)
    
    exitApp()
















def makeContentArr(page):
    soup = bs(page.text, "html.parser")
    elements = soup.select('.se-module.se-module-text')
    sentenceEndArr = ['요','죠','다','용']
    
    allStr = []
    for ele in elements:
        p = re.compile('[\uAC00-\uD7A30-9a-zA-Z\s]+')
        chkResult = p.findall(str(ele))
        allStr = allStr + chkResult

    p_str = re.compile(r'[a-zA-Z0-9,|\n]+')
    for i in range(1, len(allStr)):
        for j, strin in enumerate(allStr):
            getStr = p_str.search(strin)
            if getStr is not None:
                allStr.pop(j)
                break

            
    allStr = "".join(allStr)
    allStr = allStr.replace('   ', ' ')
    allStr = allStr.replace('   ', ' ')
    allStr = allStr.replace('   ', ' ')
    allStr = allStr.replace('  ', ' ')
    allStr = allStr.replace('  ', ' ')
    allStr = allStr.replace('  ', ' ')
    
    resetStrArr = allStr.split(' ')
    tempArr = []
    for i in range(0, len(resetStrArr)):
        for j, strin in enumerate(resetStrArr):
            if(strin != ''):
                if strin[-1] in sentenceEndArr:
                    chkArr = resetStrArr[0:j + 1]
                    tempArr.append(chkArr)
                    del resetStrArr[0:j + 1]
                    break
    
    return tempArr

def changeIp():
    try:
        os.system('adb server start')
        client = AdbClient(host="127.0.0.1", port=5037)
        device = client.devices()  # 디바이스 1개
        ondevice = device[0]
        ondevice.shell("input keyevent KEYCODE_POWER")
        ondevice.shell("svc data disable")
        ondevice.shell("settings put global airplane_mode_on 1")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

        ondevice.shell("svc data enable")
        ondevice.shell("settings put global airplane_mode_on 0")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
        time.sleep(3)
        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("https://api.ip.pe.kr/json/").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    except:

        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("https://api.ip.pe.kr/json/").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    return getIp



def searchElement(ele,driver):
    wait_float(0.3, 0.7)
    re_count = 0
    element = ""
    while True:
        re_count += 1
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
            focus_window('chrome')
            pg.press('F5')
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            pass
        

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.3, 0.7)
    return selected_element



def untilEleShow(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        try:
            btnEle = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is not None:
                return
        except:
            continue


def untilEleGone(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        
        try:
            btnEle = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is None:
                return
        except:
            return


def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)


def exitApp():
    pg.alert(text='프로그램을 종료합니다.', title='제목입니다.', button='OK')
    try:
        driver.quit()
    except:
        pass
    sys.exit(0)


def focus_window(winName):
    if winName == 'chkname':
        win_list = gw.getAllTitles()
        pg.alert(text=f"{win_list}")
    # 윈도우 타이틀에 Chrome 이 포함된 모든 윈도우 수집, 리스트로 리턴
    win = gw.getWindowsWithTitle(winName)[0]
    win.activate()  # 윈도우 활성화


BASE_DIR = Path(__file__).resolve().parent




async def getEmptyArr(setNum, exName):
    getArr = []
    asyncio.gather(*[busyFunc(i, getArr, exName) for i in range(1, setNum)])
    return getArr


async def busyFunc(i, getArr, exName):
    getTime = exName.cell(i, 4).value

    try:
        if getTime is None:
            getTime = datetime.now().date()
        if isinstance(getTime, datetime):
            getTime = getTime.date()
        compareTime = datetime.now() - timedelta(days=3)
        if exName.cell(i, 4).value is None or getTime <= compareTime.date():
            getArr.append(i)
    except:
        pass


def getExLength(exName):
    ExLength = 0
    while True:
        ExLength += 1
        if exName.cell(ExLength, 2).value is None:
            break
    return ExLength


def getUaNum():
    with open("./etc/useragent/useragent_all.txt", "r") as f:
        fArr = f.readlines()
        fCount = len(fArr)
        uaSet = random.randrange(0, fCount)
    return uaSet

def mainToCafe():
    shs_item = searchElement('.shs_item',driver)
    for item in shs_item:
        chkCafe = item.find_element(
            by=By.CSS_SELECTOR, value='a').get_attribute('href')
        if 'cafe' in chkCafe:
            untilEleGone(item, '.shs_list')
            break
    
    myCafeGo = searchElement('.mycafe .btn_cafe_more')
    untilEleGone(myCafeGo[0], '.mycafe',driver)



    myCafeList = searchElement('.list_cafe__favorites li',driver)
    with open("./etc/cafe_info.txt", "r") as f:
        getCafeNameList = f.readlines()
        getCafeName = getCafeNameList[0]
        getCafeName = getCafeName.replace(" ", "")
    
    for onCafe in myCafeList:
        chkCafeTitle = onCafe.find_element(by=By.CSS_SELECTOR, value='.title').text
        chkCafeTitle = chkCafeTitle.replace(" ", "")

        if chkCafeTitle in getCafeName:
            untilEleGone(onCafe, '.list_cafe__favorites')
            break
        
    # 카페 진입 끝

# 


# subjectArr
def list_chunk(lst, n):
    return [lst[i:i+n] for i in range(0, len(lst), n)]
