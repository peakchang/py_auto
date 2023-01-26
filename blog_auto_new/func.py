import random
import threading
import time
from datetime import datetime, timedelta
import sys
import os
from pathlib import Path
from typing import Optional
# from pyparsing import And
import requests
from bs4 import BeautifulSoup as bs
import json
import re
import pyautogui as pg
import pyperclip
import pywinauto
import pygetwindow as gw
import clipboard as cb
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import WebDriverException

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import requests
import winsound as ws
import glob
import asyncio
import winsound as sd






# 블로그 글쓰기
def writeBlog(driver,goChk):
    navItem = searchElement('.nav_item', driver)
    for mitem in navItem:
        if mitem.text == '블로그':
            mitem.click()
            break
    
    dirList = os.listdir(f"{os.getcwd()}\\content")
    
    lastForIdx = len(dirList)
    addHour = 0
    for idx,dir in enumerate(dirList):
        
        if goChk == 0:
            addHour += 1
        else:
            if idx != 0:
                addHour += 1
        menu_my_blog = searchElement('.menu_my_blog .item', driver)
        menu_my_blog[1].click()
        
        driver.switch_to.window(driver.window_handles[1])
        
        driver.switch_to.frame('mainFrame')
        
        focus_window('블로그')
        writeArea = searchElement('.se-component-content',driver)
        wait_float(2.5,3.5)
        
        try:
            preWritePopup = driver.find_element(by=By.CSS_SELECTOR, value=".se-popup-button-cancel")
            preWritePopup.click()
            wait_float(0.5,0.9)
            
        except:
            pass
        
        subject_area = writeArea[0].find_element(by=By.CSS_SELECTOR, value=".se-section-documentTitle")
        subject_area.click()
        
        try:
            with open(f'./content/{dir}/content.txt', 'rt', encoding='UTF8') as f:
                getLines = f.readlines()
        except:
            with open(f'./content/{dir}/content.txt', 'r') as f:
                getLines = f.readlines()
                
        if len(getLines) < 5:
            continue
                
                
                
        
            
        for i, getline in enumerate(getLines):
            focus_window('블로그')
            getline = getline.replace('\n', '')
            chkImg = getline.split('|')
            if chkImg[0] == 'img_line':
                nowPath = os.getcwd()
                
                try:
                    if chkImg[1]:
                        pass
                    else:
                        continue
                except:
                    continue
                img_btn = searchElement('.se-image-toolbar-button', driver)
                img_btn[0].click()
                wait_float(1.5,2.3)
                
                imagePath = nowPath + f"\content\{dir}"
                wait_float(1.5, 2.2)
                pyperclip.copy(imagePath)
                wait_float(0.5, 0.9)
                pg.hotkey('ctrl','v')
                wait_float(0.5, 0.9)
                pg.press('enter')
                
                wait_float(0.9, 1.6)
                pyperclip.copy(chkImg[1])
                wait_float(0.5, 0.9)
                pg.hotkey('ctrl','v')
                wait_float(0.5, 0.9)
                pg.press('enter')
                wait_float(3.5,4.5)
                continue
            
            if i == 0:
                writeArea[0].click()
                keyboard.write(text=getline, delay=0.05)
                wait_float(1.2,2.8)
            elif i == 1:
                writeArea[1].click()
                while True:
                    openAlignBoxBtn = searchElement('.se-align-left-toolbar-button',driver)
                    wait_float(0.5,1.2)
                    openAlignBoxBtn[0].click()
                    try:
                        alignBtnList = driver.find_elements(by=By.CSS_SELECTOR, value=".se-toolbar-option.se-toolbar-option-align button")
                        wait_float(0.5,1.2)
                        alignBtnList[1].click()
                        wait_float(0.5,1.2)
                        break
                    except:
                        pass
                keyboard.write(text=getline, delay=0.05)
                wait_float(0.5,0.9)
                pg.press('enter')
                wait_float(0.5,0.9)
            elif getline == 'enter':
                wait_float(0.5,0.9)
                pg.press('enter')
                wait_float(0.5,0.9)
            else:
                keyboard.write(text=getline, delay=0.05)
                wait_float(0.5,0.9)
                pg.press('enter')
                wait_float(0.5,0.9)
                
        try:
            helpCloseBtn = driver.find_element(by=By.CSS_SELECTOR, value=".se-help-panel-close-button")
            helpCloseBtn.click()
        except:
            pass
        
        
        
                
        
        publichBtn = searchElement('.publish_btn__Y5mLP',driver)
        publichBtn[0].click()
        wait_float(1.5,2.5)
        tagArea = searchElement('.tag_textarea__iAnXk',driver)
        tagArea[0].click()
        
        try:
            try:
                with open(f'./content/{dir}/tag_list.txt', 'rt', encoding='UTF8') as tagr:
                    tagList = tagr.readlines()
            except:
                with open(f'./content/{dir}/tag_list.txt', 'r') as tagr:
                    tagList = tagr.readlines()
        except:
            tagList = []
            pass
        
        if tagList is not [] and len(tagList) < 7:
            for tag in tagList:
                writeTag = tag.replace('\n', '')
                keyboard.write(text=writeTag, delay=0.05)
                wait_float(0.5,1.2)
                pg.press('enter')
        
        if goChk == 1:
            chkVal = pg.confirm(text='글을 지금 바로 등록하시겠습니까?', buttons=['now','reserve'])
            if chkVal == 'now':
                confirmBtn = searchElement('.confirm_btn__Dv9du',driver)
                confirmBtn[0].click()
            else:
                wait_float(0.5,0.9)
                reserveRadio = searchElement('.radio_label__wZWth',driver)
                reserveRadio[5].click()
                wait_float(0.5,0.9)
                now = datetime.now()
                setHour = now.hour + addHour
                setHourStr = str(setHour)
                planService = Select(driver.find_element(by=By.CSS_SELECTOR, value=f'.hour_option__XigHn'))
                planService.select_by_visible_text(setHourStr)
                pg.alert('예약 조건 확인 후 발행하고 엔터!!! (글발행 자동 클릭)')
                confirmBtn = searchElement('.confirm_btn__Dv9du',driver)
                confirmBtn[0].click()
        else:
            if idx == 0:
                confirmBtn = searchElement('.confirm_btn__Dv9du',driver)
                confirmBtn[0].click()
                wait_float(3.5,5.5)
            else:
                wait_float(0.5,0.9)
                reserveRadio = searchElement('.radio_label__wZWth',driver)
                reserveRadio[5].click()
                wait_float(0.5,0.9)
                now = datetime.now()
                setHour = now.hour + addHour
                setHourStr = str(setHour)
                planService = Select(driver.find_element(by=By.CSS_SELECTOR, value=f'.hour_option__XigHn'))
                planService.select_by_visible_text(setHourStr)
                confirmBtn = searchElement('.confirm_btn__Dv9du',driver)
                confirmBtn[0].click()
        
        wait_float(3.5,5.5)
        
        if idx == lastForIdx - 1:
            try:
                miniPopup = driver.find_element(by=By.CSS_SELECTOR, value="#floatingda_content")
                miniPopupClose = miniPopup.find_element(by=By.CSS_SELECTOR, value="button")
                driver.execute_script("arguments[0].scrollIntoView();", miniPopupClose)
                wait_float(0.5,1.2)
                miniPopupClose.click()
                wait_float(0.5,1.2)
            except:
                pass
            
            try:
                getUrl = searchElement('._transPosition',driver)
                getUrl[0].click()
                wait_float(1.5,2.5)
            except:
                pass
            
        driver.switch_to.default_content()
        while True:
            if len(driver.window_handles) > 1:
                driver.switch_to.window(driver.window_handles[1])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
            else:
                break
            wait_float(0.5,0.9)
    
    wait_float(2.2,2.9)
    
    try:
        closeModal = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.label_not_see._btn_no_more_show')))
        if closeModal:
            closeModal.click()
    except:
        pass
    
    goToNaverMain = searchElement('.link_naver',driver)
    goToNaverMain[0].click()
    driver.get('https://www.naver.com')
    


# 공감 순회하기
def allowListVisit(driver):
    
    navItem = searchElement('.nav_item',driver)
    for mitem in navItem:
        if mitem.text == '블로그':
            mitem.click()
            break
    
    menu_my_blog = searchElement('.menu_my_blog .item',driver)
    menu_my_blog[0].click()
    
    driver.switch_to.window(driver.window_handles[1])
    driver.switch_to.frame('mainFrame')
    
    wait_float(1.5,2.5)
    
    try:
        closePopupBtn = driver.find_element(by=By.CSS_SELECTOR, value="#not_see")
        closePopupBtn.click()
    except:
        pass
    
    try:
        closePopupBtn = driver.find_element(by=By.CSS_SELECTOR, value=".popup_da_btn_area ._btn_close")
        closePopupBtn.click()
    except:
        pass
    
    getUrl = searchElement('._transPosition',driver)
    getUrl[0].click()
    
    wait_float(1.5,2.5)
    pg.press('enter')
    
    nowBlogLink = pyperclip.paste()
    nowBlogLinkSplit = nowBlogLink.split('/')
    
    openVisitListBtn = searchElement(f'#Sympathy{nowBlogLinkSplit[-1]} .bu_arr',driver)
    openVisitListBtn[0].click()
    
    
    wait_float(1.5,2.5)
    
    driver.switch_to.frame(f'sympathyFrm{nowBlogLinkSplit[-1]}')
    
    visitListWrap = searchElement('.wrap_blog2_sympathy',driver)
    try:
        visitList = visitListWrap[0].find_elements(by=By.CSS_SELECTOR, value=".nick")
    except:
        driver.close()
        wait_float(0.3,0.9)
        driver.switch_to.window(driver.window_handles[0])
        wait_float(0.3,0.9)
        driver.switch_to.default_content()
        
        goToMain = searchElement('.link_naver',driver)
        goToMain[0].click()
        
        while True:
            try:
                WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#query")))
                break
            except:
                driver.get('https://www.naver.com')
                
        return
     
    
    
    for visitCount in range(len(visitList)):
        
        driver.switch_to.default_content()
        wait_float(0.3,0.9)
        driver.switch_to.frame('mainFrame')
        wait_float(0.3,0.9)
        driver.switch_to.frame(f'sympathyFrm{nowBlogLinkSplit[-1]}')
        wait_float(0.3,0.9)
        visitList = searchElement('.wrap_blog2_sympathy .nick',driver)
        print(visitList[visitCount].text)
        visitList[visitCount].click()
        wait_float(0.7,1.2)
        driver.switch_to.window(driver.window_handles[2])
        
        try:
            wait_float(0.7,1.2)
            driver.switch_to.default_content()
            driver.switch_to.frame('mainFrame')
            
            blogMenuChk = searchElement('#blog-menu .menu1 li a',driver)
            if(len(blogMenuChk) > 1):
                blogMenuChk[1].click()
                searchElement('#blog-menu',driver)
            
            
            
            try:
                postListOpenBtn = driver.find_element(by=By.CSS_SELECTOR, value="#toplistSpanBlind")
            except:
                wait_float(0.5,1.5)
                driver.close()
                wait_float(0.3,0.9)
                driver.switch_to.window(driver.window_handles[1])
                wait_float(0.3,0.9)
                continue
            
            
            # 여기서 블로그 말고 프롤로그면 블로그 클릭하게 하기
            
            while True:
                print(postListOpenBtn.text)
                if postListOpenBtn.text == '목록닫기':
                    break
                else:
                    wait_float(0.5,1.3)
                    try:
                        postListOpenBtn.click()
                    except:
                        wait_float(2.5,3.5)
                        continue
            wait_float(1.5,2.5)
            postList = searchElement('.blog2_categorylist',driver)
            
            for getPostLinkCount in range(3):
                try:
                    getPostLink = postList[getPostLinkCount].find_element(by=By.CSS_SELECTOR, value=".ell2")
                    break
                except:
                    pass
            
            getPostLink.click()
            wait_float(1.3,2.5)
            
            try:
                gongamBtn = driver.find_element(by=By.CSS_SELECTOR, value='.u_likeit_list_btn')
                getGonggamStatus = gongamBtn.get_attribute('aria-pressed')
                print(getGonggamStatus)
                if getGonggamStatus == 'false':
                    gongamBtn = searchElement('.u_ico',driver)
                    wait_float(0.3,0.9)
                    gongamBtn[-1].click()
                    wait_float(3.2,4.5)
                else:
                    pass
            except:
                pass

            
            # gongamBtn[1].click()
            wait_float(0.5,1.5)
            driver.close()
            wait_float(0.3,0.9)
            driver.switch_to.window(driver.window_handles[1])
            wait_float(0.3,0.9)
        except:
            wait_float(0.5,1.5)
            driver.close()
            wait_float(0.3,0.9)
            driver.switch_to.window(driver.window_handles[1])
            wait_float(0.3,0.9)
        
        
    driver.close()
    wait_float(0.3,0.9)
    driver.switch_to.window(driver.window_handles[0])
    wait_float(0.3,0.9)
    driver.switch_to.default_content()
    
    goToMain = searchElement('.link_naver',driver)
    goToMain[0].click()
    
    while True:
        try:
            WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#query")))
            break
        except:
            driver.get('https://www.naver.com')
        
    
     
    
    
    
    
    





def blogReplyReady(getValList):
    
    if getValList['nlist'] == 1:
        pg.alert('아이디가 선택되지 않았습니다. 다시 실행해주세요')
        exitApp()
    
    global driver
    
    exLineNum = getValList['nlist']
    wb = load_workbook('./etc/nid.xlsx')
    ex = wb.active
    
    options = Options()
    user_data = 'C:\\Users\\pcy\\AppData\\Local\\Google\\Chrome\\User Data\\default'
    service = Service(ChromeDriverManager().install())
    options.add_argument(f"user-data-dir={user_data}")
    options.add_argument(f'--profile-directory={ex.cell(exLineNum, 3).value}')
    driver = webdriver.Chrome(service=service, chrome_options=options)
    
    driver.get('https://www.naver.com')
    loginBtn = searchElement('.sc_login',driver)
    loginBtn[0].click()
    
    # while True:
    
    searchElement('#id',driver)
    focus_window('네이버')
    wait_float(0.3,0.9)
    while True:
        
        pg.click(200,500)
        inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
        inputId.click()
        wait_float(0.3,0.9)
        cb.copy(ex.cell(exLineNum, 1).value)
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'a')
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'v')
        inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
        if inputId.get_attribute('value') != "":
            break
        
    while True:
        inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
        inputPw.click()
        wait_float(0.3,0.9)
        cb.copy(ex.cell(exLineNum, 2).value)
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'a')
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'v')
        inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
        if inputPw.get_attribute('value') != "":
            break
    
    btnLogin = searchElement('.btn_login',driver)
    btnLogin[0].click()
    
    # 블로그 링크따기
    
    while True:
        try:
            WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#query")))
            break
        except:
            driver.get('https://www.naver.com')
    
    navItem = searchElement('.nav_item',driver)
    for mitem in navItem:
        if mitem.text == '블로그':
            mitem.click()
            break
    
    menu_my_blog = searchElement('.menu_my_blog .item',driver)
    menu_my_blog[0].click()
    
    driver.switch_to.window(driver.window_handles[1])
    driver.switch_to.frame('mainFrame')
    
    wait_float(1.5,2.5)
    
    try:
        closePopupBtn = driver.find_element(by=By.CSS_SELECTOR, value="#not_see")
        closePopupBtn.click()
    except:
        pass
    
    try:
        closePopupBtn = driver.find_element(by=By.CSS_SELECTOR, value=".popup_da_btn_area ._btn_close")
        closePopupBtn.click()
    except:
        pass
    
    getUrl = searchElement('._transPosition',driver)
    getUrl[0].click()
    wait_float(1.5,2.5)
    pg.press('enter')
    driver.switch_to.default_content()
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    
    goToNaverMain = searchElement('.link_naver',driver)
    goToNaverMain[0].click()
    
    # 블로그 링크 따기 끝~~~
    
    blogReplyWork()




def blogReplyWork(driver):
    navItem = searchElement('.nav_item',driver)
    for mitem in navItem:
        if mitem.text == '카페':
            mitem.click()
            break
    
    cafeList = searchElement('.user_mycafe_info',driver)
    getInCafe = ""
    for cafeOn in cafeList:
        if "소셜공간" in cafeOn.text:
            cafeOn.click()
            getInCafe = "on"
            break
    
    if getInCafe == "":
        driver.get('https://cafe.naver.com/sens3')
    else:
        driver.switch_to.window(driver.window_handles[0])
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        
    cafeWriteBtn = searchElement('.cafe-write-btn',driver)
    
    if "가입" in cafeWriteBtn[0].text:
        pg.alert('카페 가입하기~~')
    
    workBoardLink = searchElement('#menuLink226',driver)
    workBoardLink[0].click()
    
    driver.switch_to.frame('cafe_main')
    
    
    
    # 카페에 글 작성하기
    
    workBoardWriteBtn = searchElement('#writeFormBtn',driver)
    workBoardWriteBtn[0].click()
    
    wait_float(1.5,2.5)
    driver.switch_to.window(driver.window_handles[1])
    
    
    subjectArea = searchElement('.FlexableTextArea',driver)
    subjectArea[0].click()
    
    with open('./etc/social_cafe_content.txt', 'r') as r:
        cafeContent = r.readlines()
    
    keyboard.write(text=cafeContent[0], delay=0.05)
    wait_float(0.3,0.9)
    
    contentArea = searchElement('.se-content',driver)
    contentArea[0].click()
    wait_float(0.3,0.9)
    pg.hotkey('ctrl', 'a')
    
    for i,conLine in enumerate(cafeContent):
        if i == 0:
            continue
        keyboard.write(text=conLine, delay=0.03)
        wait_float(0.5,1.5)
    pg.press('enter')
    pg.hotkey('ctrl', 'v')
    wait_float(1.5,2.5)
    
    BaseButton = searchElement('.BaseButton',driver)
    BaseButton[0].click()
    
    
    
    
    wait_float(1.5,2.5)
    driver.switch_to.frame('cafe_main')
    
    wait_float(0.5,0.9)
    driver.close()
    
    driver.switch_to.window(driver.window_handles[0])
    
    
    
    workCafeLink = pyperclip.paste()
    workCafeNum = workCafeLink.split('/')[-1]
    preNick = ""
    
    # 카페에 글 작성하기 끝~
    
    
    
    print(workCafeLink)
    
    forVal = random.randrange(6,8)
    
    ici = 0
    while True:
        
        
        ici += 1
        driver.switch_to.default_content()
        
        wait_float(0.3,0.9)
        workBoardLink = searchElement('#menuLink226',driver)
        workBoardLink[0].click()
        
        wait_float(0.3,0.9)
        
        driver.switch_to.frame('cafe_main')
        
        wait_float(0.3,0.9)
        
        articleDiff = searchElement('.article-board',driver)
        articleList = articleDiff[1].find_elements(by=By.CSS_SELECTOR, value=".td_article")
        
        
        
        if str(workCafeNum) in articleList[ici].find_element(by=By.CSS_SELECTOR, value=".board-number").text:
            forVal = forVal + 1
            continue
        
        wait_float(0.3,0.9)
        clickArticleTarget = articleList[ici].find_element(by=By.CSS_SELECTOR, value=".article")
        wait_float(0.3,0.9)
        # clickArticleTarget.click()
        untilEleShow(clickArticleTarget, '.nickname')
        
        
        nickname = searchElement('.nickname',driver)
        if preNick == nickname[0].text:
            driver.back()
            wait_float(2.1,3.7)
            continue
        else:
            preNick = nickname[0].text
        
        chkLinkTag = driver.find_elements("xpath", "//*[contains(@class, 'se-fs-')]")

        for chkLink in chkLinkTag:
            try:
                getOtherBlogLink = chkLink.find_element(by=By.CSS_SELECTOR, value="a").get_attribute('href')
                if 'blog' in str(getOtherBlogLink):
                    chkOtherBlogLink = getOtherBlogLink.split('/')
                    if len(chkOtherBlogLink) < 5:
                        forVal = forVal + 1
                        driver.back()
                        wait_float(2.1,3.7)
                        break
                    else:
                        
                        chkLink.click()
                        
                        if 'm.' in getOtherBlogLink:
                            wait_float(0.3,0.9)
                            driver.switch_to.window(driver.window_handles[1])
                            driver.switch_to.default_content()
                            gongamBtn = searchElement('.u_ico',driver)
                            wait_float(0.3,0.9)
                            gongamBtn[-1].click()
                            wait_float(2.2,2.9)
                            driver.close()
                            wait_float(0.3,0.9)
                            driver.switch_to.window(driver.window_handles[0])
                            wait_float(0.3,0.9)
                            
                        else:
                            wait_float(0.3,0.9)
                            driver.switch_to.window(driver.window_handles[1])
                            driver.switch_to.default_content()
                            wait_float(0.3,0.9)
                            driver.switch_to.frame('mainFrame')
                            wait_float(0.3,0.9)
                            gongamBtn = searchElement('.u_ico',driver)
                            gongamBtn[1].click()
                            wait_float(2.2,2.9)
                            driver.switch_to.default_content()
                            wait_float(0.3,0.9)
                            driver.close()
                            wait_float(0.3,0.9)
                            driver.switch_to.window(driver.window_handles[0])
                            wait_float(0.3,0.9)
                        
                        driver.switch_to.frame('cafe_main')
                        replyArea = searchElement('.comment_inbox_text',driver)
                        replyArea[0].click()
                        
                        for i,conLine in enumerate(cafeContent):
                            if i == 0:
                                continue
                            keyboard.write(text=conLine, delay=0.03)
                            wait_float(0.5,1.5)
                        pg.press('enter')
                        pg.hotkey('ctrl', 'v')
                        wait_float(1.5,2.5)
                        
                        replySuccessBtn = searchElement('.btn_register',driver)
                        driver.execute_script("arguments[0].scrollIntoView();", replySuccessBtn[0])
                        replySuccessBtn[0].click()
                        wait_float(5.5,7.5)
                        
                        driver.back()
                        break
            except:
                pass
        if ici >= forVal:
            break









# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>함수 시작염


def changeIp():
    try:
        os.system('adb server start')
        client = AdbClient(host="127.0.0.1", port=5037)
        device = client.devices()  # 디바이스 1개
        ondevice = device[0]
        ondevice.shell("input keyevent KEYCODE_POWER")
        ondevice.shell("svc data disable")
        ondevice.shell("settings put global airplane_mode_on 1")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

        ondevice.shell("svc data enable")
        ondevice.shell("settings put global airplane_mode_on 0")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
        time.sleep(3)
        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("https://api.ip.pe.kr/json/").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    except:

        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("https://api.ip.pe.kr/json/").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    return getIp



def searchElement(ele,driver):
    wait_float(0.3, 0.7)
    re_count = 0
    element = ""
    while True:
        re_count += 1
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
            
            pg.press('F5')
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            pass
        

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.3, 0.7)
    return selected_element



def untilEleShow(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        try:
            btnEle = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is not None:
                return
        except:
            continue


def untilEleGone(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        
        try:
            btnEle = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is None:
                return
        except:
            return


def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)


def exitApp():
    pg.alert(text='프로그램을 종료합니다.', title='제목입니다.', button='OK')
    try:
        driver.quit()
    except:
        pass
    sys.exit(0)


def focus_window(winName):
    while True:
        activeName = str(pg.getActiveWindow())
        wait_float(0.3,0.9)
        if winName in activeName:
            return
        winList = pg.getAllWindows()
        # pg.alert(winList)
        wait_float(0.3,0.9)
        for win in winList:
            if winName in win.title:
                pywinauto.application.Application().connect(handle=win._hWnd).top_window().set_focus()
                win.activate()
                wait_float(0.7,1.3)
                break


BASE_DIR = Path(__file__).resolve().parent




async def getEmptyArr(setNum, exName):
    getArr = []
    asyncio.gather(*[busyFunc(i, getArr, exName) for i in range(1, setNum)])
    return getArr


async def busyFunc(i, getArr, exName):
    getTime = exName.cell(i, 4).value

    try:
        if getTime is None:
            getTime = datetime.now().date()
        if isinstance(getTime, datetime):
            getTime = getTime.date()
        compareTime = datetime.now() - timedelta(days=3)
        if exName.cell(i, 4).value is None or getTime <= compareTime.date():
            getArr.append(i)
    except:
        pass


def getExLength(exName):
    ExLength = 0
    while True:
        ExLength += 1
        if exName.cell(ExLength, 2).value is None:
            break
    return ExLength


def getUaNum():
    with open("./etc/useragent/useragent_all.txt", "r") as f:
        fArr = f.readlines()
        fCount = len(fArr)
        uaSet = random.randrange(0, fCount)
    return uaSet

def mainToCafe():
    shs_item = searchElement('.shs_item',driver)
    for item in shs_item:
        chkCafe = item.find_element(
            by=By.CSS_SELECTOR, value='a').get_attribute('href')
        if 'cafe' in chkCafe:
            untilEleGone(item, '.shs_list')
            break
    
    myCafeGo = searchElement('.mycafe .btn_cafe_more')
    untilEleGone(myCafeGo[0], '.mycafe',driver)



    myCafeList = searchElement('.list_cafe__favorites li',driver)
    with open("./etc/cafe_info.txt", "r") as f:
        getCafeNameList = f.readlines()
        getCafeName = getCafeNameList[0]
        getCafeName = getCafeName.replace(" ", "")
    
    for onCafe in myCafeList:
        chkCafeTitle = onCafe.find_element(by=By.CSS_SELECTOR, value='.title').text
        chkCafeTitle = chkCafeTitle.replace(" ", "")

        if chkCafeTitle in getCafeName:
            untilEleGone(onCafe, '.list_cafe__favorites')
            break
        
    # 카페 진입 끝

# 


# subjectArr
def list_chunk(lst, n):
    return [lst[i:i+n] for i in range(0, len(lst), n)]
